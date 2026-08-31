from __future__ import annotations

import base64
import hashlib
import json
import os
from types import SimpleNamespace
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from storage_intelligence import IntelligenceEngine, generate_accounts
from storage_intelligence.analytics import risk_score, security_findings
from storage_intelligence.connectors import (
    ConnectorContext,
    ConnectorDisabledError,
    ResourceGraphConnector,
)
from storage_intelligence.hierarchy import SUBSCRIPTIONS, TENANTS
from storage_intelligence.synthetic import dataset_fingerprint

QUESTIONS = [
    "Which accounts grew abnormally last week?",
    "Which Storage Accounts should I worry about, and why?",
    "Which accounts are negatively impacting Databricks?",
    "How much can we save by tiering cold data?",
    "Why did storage costs increase 22% last month?",
    "What should move to Cool, Cold, or Archive?",
    "Where are we wasting money?",
    "When will we need more capacity?",
    "Which business unit drives growth?",
    "Which subscriptions are outliers?",
    "Compare risk across management groups and subsidiaries.",
    "Which SAP-linked storage accounts should be reviewed first?",
    "Which Azure Data Factory-linked accounts have the highest operational pressure?",
    "Which stale accounts also have high cost or operational risk?",
    "Which accounts combine rapid growth with weak lifecycle coverage?",
    "Compare storage risk, growth, and cost across Dev, QA, Perf, and Prod environments.",
    "Which storage accounts still use SAS tokens or shared keys?",
    "Which publicly accessible accounts have no private endpoint?",
    "Which accounts have no service-principal-based access enabled?",
    "Which NSG or ASG-linked accounts still expose public access?",
    "Which accounts lack GRS or GZRS resilience?",
    "Which defunct projects still retain storage accounts?",
    "Which projects have not accessed their storage in more than 180 days?",
    "Which SFTP-enabled storage accounts have public access or no private endpoint?",
    "Which accounts storing Application Insights data have weak security or resilience?",
    "Which accounts will cross a capacity threshold first?",
    "Which lifecycle policies are missing?",
    "Where would Archive create high rehydration or early-deletion risk?",
    "Which accounts have high transaction cost relative to stored TB?",
    "Which containers show small-file patterns that hurt Databricks?",
    "Which Databricks external locations drive storage IO?",
    "Which accounts have stale or missing inventory reports?",
    "Which accounts have versioning, snapshots, or soft-delete retention driving growth?",
    "Which replication choices appear over-provisioned?",
    "What changed since the previous weekly review?",
    "What are the top five actions by savings, risk reduction, and implementation effort?",
    "Show a what-if comparison for 10%, 25%, and 50% tiering adoption.",
    "Which findings are based on incomplete data?",
]


def _returned_account_items(value):
    items = []
    if isinstance(value, dict):
        if value.get("account_id"):
            items.append(value)
        for nested in value.values():
            items.extend(_returned_account_items(nested))
    elif isinstance(value, list):
        for nested in value:
            items.extend(_returned_account_items(nested))
    return items


@pytest.fixture(scope="module")
def accounts():
    return generate_accounts()


def test_synthetic_estate_is_large_and_deterministic(accounts):
    assert len(accounts) == 2500
    assert len({row["account_id"] for row in accounts}) == 2500
    assert dataset_fingerprint(accounts) == dataset_fingerprint(generate_accounts())
    assert len({row["subscription"] for row in accounts}) == 339
    assert {row["environment"] for row in accounts} == {"Dev", "QA", "Perf", "Prod"}
    assert len({row["tenant_id"] for row in accounts}) == 3
    assert len({row["management_group"] for row in accounts}) >= 4
    assert len({row["subsidiary"] for row in accounts}) >= 5
    assert all(row["subsidiary"] == row["business_unit"] for row in accounts)
    assert any(row["databricks_workspace"] for row in accounts)
    assert any(row["fabric_lakehouse"] for row in accounts)
    assert any(row["sap_system"] for row in accounts)
    assert any(row["azure_data_factory"] for row in accounts)
    assert any(row["uses_sas_keys"] for row in accounts)
    assert any(row["public_network_access"] for row in accounts)
    assert any(not row["private_endpoint_enabled"] for row in accounts)
    assert any(not row["service_principal_access_enabled"] for row in accounts)
    assert any(row["network_security_group"] or row["application_security_group"] for row in accounts)
    assert any(row["project_defunct"] for row in accounts)
    assert any(row["sftp_enabled"] for row in accounts)
    assert any(row["application_insights_resource"] for row in accounts)
    assert any(row["azure_function_app"] for row in accounts)
    assert any(row["log_analytics_workspace"] for row in accounts)
    assert any(row["azure_function_app"] and row["log_analytics_workspace"] for row in accounts)
    assert any(row["managed_identity_enabled"] for row in accounts)
    assert any(not row["managed_identity_enabled"] for row in accounts)
    assert all(not row["sftp_enabled"] or row["hns_enabled"] for row in accounts)
    assert all(row["project_name"] and row["tag_business_unit"] and row["last_accessed_date"] for row in accounts)


def test_subscription_catalog_has_339_environment_mappings():
    assert len(SUBSCRIPTIONS) == 339
    assert len({item["id"] for item in SUBSCRIPTIONS}) == 339
    assert len({item["name"] for item in SUBSCRIPTIONS}) == 339
    assert {item["environment"] for item in SUBSCRIPTIONS} == {"Dev", "QA", "Perf", "Prod"}
    assert all(item["tenant_id"] and item["management_group"] and item["subsidiary"] for item in SUBSCRIPTIONS)
    subsidiaries = {item["subsidiary"] for item in SUBSCRIPTIONS}
    assert {"Avidunixuser Research", "Avidunixuser OSS"} <= subsidiaries


@pytest.mark.parametrize("question", QUESTIONS)
def test_query_catalog_contract(accounts, question):
    result = IntelligenceEngine(accounts).answer(question)
    assert result["scope"]["account_count"] == 2500
    assert result["timestamp"]
    assert result["data_as_of"]
    assert result["evidence"]
    assert result["assumptions"]
    assert result["confidence"]["level"] in {"low", "medium", "high"}
    assert result["tool"]
    assert result["data"] is not None
    account_items = _returned_account_items(result["data"])
    if account_items:
        assert all(item.get("reason") for item in account_items)
        returned_ids = list(dict.fromkeys(item["account_id"] for item in account_items))
        assert [item["account_id"] for item in result["account_reasons"]] == returned_ids
        assert [item["id"] for item in result["evidence"]] == returned_ids
        assert all(item.get("reason") for item in result["evidence"])


def test_risk_evidence_matches_every_ranked_account(accounts):
    result = IntelligenceEngine(accounts).answer("Which Storage Accounts should I worry about, and why?")
    ranked_ids = [item["account_id"] for item in result["data"]]

    assert len(ranked_ids) > 3
    assert [item["id"] for item in result["evidence"]] == ranked_ids
    assert [item["account_id"] for item in result["account_reasons"]] == ranked_ids
    assert all("Overall risk is" in item["reason"] for item in result["data"])


def test_security_and_governance_risk_factors_are_transparent(accounts):
    row = dict(accounts[0])
    row.update(
        {
            "uses_sas_keys": True,
            "shared_key_access_enabled": True,
            "public_network_access": True,
            "blob_public_access_enabled": True,
            "private_endpoint_enabled": False,
            "service_principal_access_enabled": False,
            "network_security_group": "nsg-risk",
            "application_security_group": "asg-risk",
            "replication": "LRS",
            "project_defunct": True,
            "last_accessed_date": "2024-01-01",
        }
    )
    scored = risk_score(row)

    assert scored["components"]["security"] == 100
    assert scored["components"]["governance"] >= 90
    assert len(scored["risk_factors"]) >= 8
    assert any("SAS" in factor for factor in scored["risk_factors"])
    assert any("Public network" in factor for factor in scored["risk_factors"])
    assert any("private endpoint" in factor for factor in scored["risk_factors"])
    assert any("GRS/GZRS" in factor for factor in scored["risk_factors"])
    assert any("service-principal" in factor for factor in scored["risk_factors"])
    assert security_findings([row])[0]["account_id"] == row["account_id"]


def test_platform_specific_question_routing(accounts):
    engine = IntelligenceEngine(accounts)
    sap = engine.answer("Which SAP-linked storage accounts should be reviewed first?")
    adf = engine.answer("Which Azure Data Factory-linked accounts have the highest operational pressure?")

    assert sap["tool"] == "risk.sap"
    assert adf["tool"] == "risk.data_factory"
    assert all(item["sap_system"] for item in sap["data"])
    assert all(item["azure_data_factory"] for item in adf["data"])


def test_security_governance_and_resilience_question_routing(accounts):
    engine = IntelligenceEngine(accounts)
    security = engine.answer("Which publicly accessible accounts have no private endpoint?")
    governance = engine.answer("Which defunct projects still retain storage accounts?")
    resilience = engine.answer("Which accounts lack GRS or GZRS resilience?")

    assert security["tool"] == "risk.security_posture"
    assert governance["tool"] == "risk.project_governance"
    assert resilience["tool"] == "risk.resilience"
    assert all(item["components"]["security"] > 0 for item in security["data"])
    assert all(item["components"]["governance"] > 0 for item in governance["data"])
    assert all(
        "does not provide GRS/GZRS" in " ".join(item["risk_factors"])
        for item in resilience["data"]
    )


def test_spanish_question_intents_route_to_deterministic_tools(accounts):
    engine = IntelligenceEngine(accounts)

    security = engine.answer(
        "¿Qué cuentas con acceso público no tienen un punto de conexión privado?"
    )
    savings = engine.answer(
        "¿Cuánto podemos ahorrar al cambiar de nivel los datos fríos?"
    )
    governance = engine.answer(
        "¿Qué proyecto obsoleto conserva cuentas y cuál fue su último acceso?"
    )

    assert security["tool"] == "risk.security_posture"
    assert savings["tool"] == "cost.tier_savings"
    assert governance["tool"] == "risk.project_governance"


def test_sftp_and_application_insights_question_routing(accounts):
    engine = IntelligenceEngine(accounts)
    sftp = engine.answer("Which SFTP-enabled storage accounts have public access or no private endpoint?")
    app_insights = engine.answer(
        "Which accounts storing Application Insights data have weak security or resilience?"
    )

    assert sftp["tool"] == "risk.sftp"
    assert app_insights["tool"] == "risk.application_insights_storage"
    assert all(item["sftp_enabled"] for item in sftp["data"])
    assert all(item["application_insights_resource"] for item in app_insights["data"])


def test_filters_are_enforced(accounts):
    result = IntelligenceEngine(accounts).answer(
        "Which accounts are risky?",
        {
            "tenant_id": TENANTS[0]["id"],
            "management_group": SUBSCRIPTIONS[0]["management_group"],
            "subsidiary": SUBSCRIPTIONS[0]["subsidiary"],
            "subscription": "platform-prod",
            "environment": "Prod",
            "region": "eastus2",
        },
    )
    assert 0 < result["scope"]["account_count"] < 2500
    assert result["scope"]["filters"]["subscription"] == "platform-prod"
    assert result["scope"]["filters"]["subsidiary"] == SUBSCRIPTIONS[0]["subsidiary"]
    assert result["scope"]["filters"]["environment"] == "Prod"


def test_connectors_are_disabled_by_default(monkeypatch):
    monkeypatch.delenv("ENABLE_RESOURCE_GRAPH", raising=False)
    connector = ResourceGraphConnector(
        ConnectorContext(source="resource-graph", enabled=False),
        ["00000000-0000-0000-0000-000000000000"],
    )
    with pytest.raises(ConnectorDisabledError):
        list(connector.collect())


def test_web_api_and_auth_boundary(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import app

    client = TestClient(app)
    assert client.get("/healthz").status_code == 200
    ready = client.get("/readyz")
    assert ready.status_code == 200
    assert ready.json()["dataset_accounts"] == 2500
    portfolio = client.get("/api/portfolio")
    assert portfolio.status_code == 200
    portfolio_body = portfolio.json()
    assert portfolio_body["summary"]["account_count"] == 2500
    assert portfolio_body["hierarchy"] == {
        "tenant_count": 3,
        "management_group_count": 4,
        "subsidiary_count": 5,
        "subscription_count": 339,
        "environment_count": 4,
    }
    assert len(portfolio_body["filters"]["tenant_ids"]) == 3
    assert len(portfolio_body["filters"]["management_groups"]) == 4
    assert len(portfolio_body["filters"]["subsidiaries"]) == 5
    assert len(portfolio_body["filters"]["subscriptions"]) == 339
    assert portfolio_body["filters"]["environments"] == ["Dev", "Perf", "Prod", "QA"]
    assert portfolio_body["risk_threshold"] == 20.0
    assert portfolio_body["summary"]["sas_key_accounts"] > 0
    assert portfolio_body["summary"]["public_access_accounts"] > 0
    assert portfolio_body["summary"]["missing_private_endpoint_accounts"] > 0
    assert portfolio_body["summary"]["missing_service_principal_access_accounts"] > 0
    assert portfolio_body["summary"]["managed_identity_accounts"] > 0
    assert portfolio_body["summary"]["non_geo_redundant_accounts"] > 0
    assert portfolio_body["summary"]["nsg_asg_linked_accounts"] > 0
    assert portfolio_body["summary"]["defunct_project_accounts"] > 0
    assert portfolio_body["summary"]["sftp_enabled_accounts"] > 0
    assert portfolio_body["summary"]["application_insights_accounts"] > 0
    assert len(portfolio_body["risks"]) == portfolio_body["summary"]["at_risk_accounts"]
    assert len(portfolio_body["risks"]) > 8
    assert all(item["score"] >= portfolio_body["risk_threshold"] for item in portfolio_body["risks"])
    assert all("security" in item["components"] and "governance" in item["components"] for item in portfolio_body["risks"])
    assert any(item["risk_factors"] for item in portfolio_body["risks"])
    assert portfolio_body["risks"] == sorted(
        portfolio_body["risks"],
        key=lambda item: item["score"],
        reverse=True,
    )
    assert len(portfolio_body["platform_accounts"]) == 2500
    assert len({item["account_id"] for item in portfolio_body["platform_accounts"]}) == 2500
    assert any(item["databricks_workspace"] for item in portfolio_body["platform_accounts"])
    assert any(item["fabric_lakehouse"] for item in portfolio_body["platform_accounts"])
    assert any(item["sap_system"] for item in portfolio_body["platform_accounts"])
    assert any(item["azure_data_factory"] for item in portfolio_body["platform_accounts"])
    assert any(item["sftp_enabled"] for item in portfolio_body["platform_accounts"])
    assert any(item["application_insights_resource"] for item in portfolio_body["platform_accounts"])
    assert any(item["azure_function_app"] for item in portfolio_body["platform_accounts"])
    assert any(item["log_analytics_workspace"] for item in portfolio_body["platform_accounts"])
    assert any(item["managed_identity_enabled"] for item in portfolio_body["platform_accounts"])
    assert any(not item["managed_identity_enabled"] for item in portfolio_body["platform_accounts"])
    assert all(
        {"business_unit", "project_name", "tag_business_unit", "project_defunct", "business_criticality", "hns_enabled"}
        <= item.keys()
        for item in portfolio_body["platform_accounts"]
    )
    answer = client.post("/api/query", json={"question": QUESTIONS[0], "filters": {}})
    assert answer.status_code == 200
    assert answer.json()["evidence"]


def test_agent_query_returns_foundry_usage(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web import app as web_app

    monkeypatch.setattr(
        web_app,
        "invoke_agent",
        lambda question, filters: {
            "conversation_id": "conversation-1",
            "answer": "Foundry answer",
            "model": "gpt-5.4-mini",
            "usage": {
                "input_tokens": 1200,
                "output_tokens": 300,
                "total_tokens": 1500,
                "context_used_tokens": 1200,
                "context_window": 400000,
            },
        },
    )
    response = TestClient(web_app.app).post(
        "/api/agent/query",
        json={"question": QUESTIONS[0], "filters": {"environment": "Prod"}},
    )

    assert response.status_code == 200
    body = response.json()
    assert body["answer"] == "Foundry answer"
    assert body["evidence"]
    assert body["agent"] == {
        "conversation_id": "conversation-1",
        "model": "gpt-5.4-mini",
        "usage": {
            "input_tokens": 1200,
            "output_tokens": 300,
            "total_tokens": 1500,
            "context_used_tokens": 1200,
            "context_window": 400000,
        },
    }


def test_foundry_response_payload_reports_model_tokens_and_context(monkeypatch):
    monkeypatch.setenv("AZURE_AI_MODEL_CONTEXT_WINDOW", "400000")
    from agent.invoke import _response_payload

    payload = _response_payload(
        SimpleNamespace(
            output_text="Foundry answer",
            usage=SimpleNamespace(input_tokens=800, output_tokens=200, total_tokens=1000),
        ),
        conversation_id="conversation-2",
        model="gpt-5.4-mini",
    )

    assert payload["model"] == "gpt-5.4-mini"
    assert payload["usage"] == {
        "input_tokens": 800,
        "output_tokens": 200,
        "total_tokens": 1000,
        "context_used_tokens": 800,
        "context_window": 400000,
    }


@pytest.mark.parametrize(
    ("factor", "summary_key"),
    [
        ("sas-key", "sas_key_accounts"),
        ("public-access", "public_access_accounts"),
        ("no-private-endpoint", "missing_private_endpoint_accounts"),
        ("no-service-principal", "missing_service_principal_access_accounts"),
        ("managed-identity", "managed_identity_accounts"),
        ("no-grs-gzrs", "non_geo_redundant_accounts"),
        ("nsg-asg-linked", "nsg_asg_linked_accounts"),
        ("defunct-projects", "defunct_project_accounts"),
        ("sftp-enabled", "sftp_enabled_accounts"),
        ("app-insights-data", "application_insights_accounts"),
        ("stale-accounts", "stale_accounts"),
        ("missing-lifecycle", "missing_lifecycle_policy"),
    ],
)
def test_posture_drilldown_matches_overview_tiles(monkeypatch, factor, summary_key):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import app

    client = TestClient(app)
    portfolio = client.get("/api/portfolio").json()
    data_health = client.get("/api/data-health").json()
    response = client.get(f"/api/posture/{factor}")

    assert response.status_code == 200
    body = response.json()
    summary = (
        data_health["summary"]
        if factor in {"stale-accounts", "missing-lifecycle"}
        else portfolio["summary"]
    )
    assert body["count"] == summary[summary_key]
    assert len(body["accounts"]) == body["count"]
    assert body["accounts"] == sorted(
        body["accounts"],
        key=lambda item: (-item["score"], item["name"]),
    )
    assert all(item["detail"] and item["account_id"] for item in body["accounts"])
    assert body["scope"]["account_count"] == 2500


def test_posture_drilldown_honors_filters_and_allowlist(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import app

    client = TestClient(app)
    response = client.get("/api/posture/public-access", params={"environment": "Prod"})
    assert response.status_code == 200
    body = response.json()
    assert body["scope"]["filters"] == {"environment": "Prod"}
    assert body["scope"]["environment_count"] == 1
    assert all(item["environment"] == "Prod" for item in body["accounts"])
    assert client.get("/api/posture/arbitrary-factor").status_code == 404


def test_add_storage_account_to_pilot_inventory(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import ACCOUNTS, app

    client = TestClient(app)
    before = len(ACCOUNTS)
    payload = {
        "name": "stmanualpilot01",
        "tenant_id": TENANTS[0]["id"],
        "management_group": SUBSCRIPTIONS[0]["management_group"],
        "subscription": "platform-prod",
        "environment": "Prod",
        "subsidiary": SUBSCRIPTIONS[0]["subsidiary"],
        "region": "eastus2",
        "tier": "Hot",
    }
    try:
        response = client.post("/api/accounts", json=payload)
        assert response.status_code == 201
        assert response.json()["account"]["source"] == "manual-pilot-v1"
        assert response.json()["account"]["tenant_id"] == TENANTS[0]["id"]
        assert response.json()["account"]["management_group"] == SUBSCRIPTIONS[0]["management_group"]
        assert response.json()["account"]["environment"] == "Prod"
        assert response.json()["account"]["subsidiary"] == SUBSCRIPTIONS[0]["subsidiary"]
        assert response.json()["account"]["business_unit"] == SUBSCRIPTIONS[0]["subsidiary"]
        assert response.json()["total"] == before + 1
        assert client.post("/api/accounts", json=payload).status_code == 409
        portfolio = client.get("/api/portfolio")
        assert portfolio.json()["summary"]["account_count"] == before + 1
    finally:
        ACCOUNTS[:] = [row for row in ACCOUNTS if row["name"] != payload["name"]]


def test_bulk_csv_account_import(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import ACCOUNTS, app

    names = {"stbulkcsv01", "stbulkcsv02"}
    content = (
        "Storage account name,Tenant ID,Management group,Subscription,Environment,Subsidiary,Region,Access tier\n"
        f"stbulkcsv01,{TENANTS[0]['id']},{SUBSCRIPTIONS[0]['management_group']},platform-prod,Prod,{SUBSCRIPTIONS[0]['subsidiary']},eastus2,Hot\n"
        f"stbulkcsv02,{TENANTS[0]['id']},{SUBSCRIPTIONS[1]['management_group']},data-prod,Prod,{SUBSCRIPTIONS[1]['subsidiary']},westeurope,Cool\n"
    ).encode()
    try:
        response = TestClient(app).post(
            "/api/accounts/import",
            files={"spreadsheet": ("accounts.csv", content, "text/csv")},
        )
        assert response.status_code == 201
        assert response.json()["imported"] == 2
        assert {row["name"] for row in ACCOUNTS}.issuperset(names)
        assert all(row["source"] == "spreadsheet-pilot-v1" for row in ACCOUNTS if row["name"] in names)
    finally:
        ACCOUNTS[:] = [row for row in ACCOUNTS if row["name"] not in names]


def test_bulk_xlsx_account_import_is_atomic(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import ACCOUNTS, app

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "tenant_id", "management_group", "subscription", "environment", "subsidiary", "region", "tier"])
    sheet.append(["stbulkxlsx01", TENANTS[0]["id"], SUBSCRIPTIONS[0]["management_group"], "platform-prod", "Prod", SUBSCRIPTIONS[0]["subsidiary"], "eastus2", "Hot"])
    sheet.append(["stbulkxlsx02", TENANTS[0]["id"], SUBSCRIPTIONS[1]["management_group"], "data-prod", "Prod", SUBSCRIPTIONS[1]["subsidiary"], "westeurope", "invalid-tier"])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()

    before = len(ACCOUNTS)
    response = TestClient(app).post(
        "/api/accounts/import",
        files={
            "spreadsheet": (
                "accounts.xlsx",
                payload.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert response.status_code == 422
    assert len(ACCOUNTS) == before
    assert not any(row["name"] in {"stbulkxlsx01", "stbulkxlsx02"} for row in ACCOUNTS)


def test_bulk_xlsx_account_import(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import ACCOUNTS, app

    names = {"stxlsxvalid01", "stxlsxvalid02"}
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["name", "tenant_id", "management_group", "subscription", "environment", "business_unit", "region", "tier"])
    sheet.append(["stxlsxvalid01", TENANTS[0]["id"], SUBSCRIPTIONS[0]["management_group"], "platform-prod", "Prod", SUBSCRIPTIONS[0]["subsidiary"], "eastus2", "Hot"])
    sheet.append(["stxlsxvalid02", TENANTS[0]["id"], SUBSCRIPTIONS[1]["management_group"], "data-prod", "Prod", SUBSCRIPTIONS[1]["subsidiary"], "westeurope", "Cool"])
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()

    try:
        response = TestClient(app).post(
            "/api/accounts/import",
            files={
                "spreadsheet": (
                    "accounts.xlsx",
                    payload.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 201
        assert response.json()["imported"] == 2
        assert {row["name"] for row in ACCOUNTS}.issuperset(names)
    finally:
        ACCOUNTS[:] = [row for row in ACCOUNTS if row["name"] not in names]


def test_bulk_xlsx_import_persists_full_airgap_record(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    import web.app as web_app

    account_name = "stairgapcosmos01"
    account_id = (
        "/subscriptions/11111111-1111-1111-1111-111111111111/resourceGroups/rg-airgap/"
        f"providers/Microsoft.Storage/storageAccounts/{account_name}"
    )
    headers = [
        "account_id",
        "name",
        "tenant_id",
        "subscription_id",
        "subscription_name",
        "environment",
        "management_group",
        "subsidiary",
        "business_unit",
        "region",
        "tier",
        "tier_assumed",
        "kind",
        "sku",
        "uses_sas_keys",
        "shared_key_access_enabled",
        "public_network_access",
        "blob_public_access_enabled",
        "private_endpoint_enabled",
        "service_principal_access_enabled",
        "managed_identity_enabled",
        "network_security_group",
        "application_security_group",
        "project_name",
        "tag_business_unit",
        "last_accessed_date",
        "project_defunct",
        "databricks_workspace",
        "fabric_lakehouse",
        "sap_system",
        "azure_data_factory",
        "hns_enabled",
        "sftp_enabled",
        "application_insights_resource",
        "azure_function_app",
        "log_analytics_workspace",
    ]
    values = [
        account_id,
        account_name,
        TENANTS[0]["id"],
        "11111111-1111-1111-1111-111111111111",
        "platform-prod",
        "Prod",
        SUBSCRIPTIONS[0]["management_group"],
        SUBSCRIPTIONS[0]["subsidiary"],
        "AIRGAP Data",
        "eastus2",
        "Hot",
        True,
        "StorageV2",
        "Standard_ZRS",
        False,
        False,
        False,
        False,
        True,
        True,
        True,
        "nsg-airgap",
        "asg-airgap",
        "project-airgap",
        "AIRGAP Data",
        "2026-08-20",
        False,
        "dbw-airgap",
        "lakehouse-airgap",
        "SAP-AIRGAP",
        "adf-airgap",
        True,
        False,
        "appi-airgap",
        "func-airgap",
        "log-airgap",
    ]
    workbook = Workbook()
    workbook.active.append(headers)
    workbook.active.append(values)
    content = BytesIO()
    workbook.save(content)
    workbook.close()
    persisted = []

    def fake_persist(accounts, *, pulled_at, trigger, source):
        assert not any(row["name"] == account_name for row in web_app.ACCOUNTS)
        assert trigger == "airgap-spreadsheet"
        assert source == "spreadsheet-pilot-v1"
        assert pulled_at == accounts[0]["data_as_of"]
        persisted.extend(accounts)
        return {
            "status": "completed",
            "upserted": len(accounts),
            "database": "storage-intelligence",
            "container": "storage-accounts",
        }

    monkeypatch.setattr(web_app, "persist_inventory_accounts", fake_persist)
    try:
        response = TestClient(web_app.app).post(
            "/api/accounts/import",
            files={
                "spreadsheet": (
                    "airgap.xlsx",
                    content.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert response.status_code == 201
        assert response.json()["persistence"]["upserted"] == 1
        assert persisted[0]["account_id"] == account_id
        assert persisted[0]["subscription_id"] == values[3]
        assert persisted[0]["subscription_name"] == "platform-prod"
        assert persisted[0]["business_unit"] == "AIRGAP Data"
        assert persisted[0]["tier_assumed"] is True
        assert persisted[0]["kind"] == "StorageV2"
        assert persisted[0]["sku"] == "Standard_ZRS"
        assert persisted[0]["uses_sas_keys"] is False
        assert persisted[0]["private_endpoint_enabled"] is True
        assert persisted[0]["project_name"] == "project-airgap"
        assert persisted[0]["last_accessed_date"] == "2026-08-20"
        assert persisted[0]["hns_enabled"] is True
        assert persisted[0]["sftp_enabled"] is False
        assert persisted[0]["log_analytics_workspace"] == "log-airgap"
        assert any(row["name"] == account_name for row in web_app.ACCOUNTS)
    finally:
        web_app.ACCOUNTS[:] = [row for row in web_app.ACCOUNTS if row["name"] != account_name]


def test_bulk_import_does_not_publish_when_cosmos_persistence_fails(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    import web.app as web_app

    account_name = "stairgapfailure01"
    content = (
        "name,tenant_id,management_group,subscription,environment,subsidiary,region,tier\n"
        f"{account_name},{TENANTS[0]['id']},{SUBSCRIPTIONS[0]['management_group']},"
        f"platform-prod,Prod,{SUBSCRIPTIONS[0]['subsidiary']},eastus2,Hot\n"
    ).encode()

    def fail_persistence(*args, **kwargs):
        raise RuntimeError("Cosmos unavailable")

    monkeypatch.setattr(web_app, "persist_inventory_accounts", fail_persistence)
    response = TestClient(web_app.app).post(
        "/api/accounts/import",
        files={"spreadsheet": ("airgap.csv", content, "text/csv")},
    )

    assert response.status_code == 503
    assert response.json()["detail"] == (
        "Spreadsheet validated but the accounts could not be persisted to Cosmos DB"
    )
    assert not any(row["name"] == account_name for row in web_app.ACCOUNTS)


def test_bulk_import_derives_cosmos_partition_from_account_id(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    import web.app as web_app

    account_name = "stairgappartition01"
    subscription_id = "22222222-2222-2222-2222-222222222222"
    account_id = (
        f"/subscriptions/{subscription_id}/resourceGroups/rg-airgap/"
        f"providers/Microsoft.Storage/storageAccounts/{account_name}"
    )
    content = (
        "account_id,name,tenant_id,management_group,subscription,environment,"
        "subsidiary,region,tier\n"
        f"{account_id},{account_name},{TENANTS[0]['id']},"
        f"{SUBSCRIPTIONS[0]['management_group']},platform-prod,Prod,"
        f"{SUBSCRIPTIONS[0]['subsidiary']},eastus2,Hot\n"
    ).encode()
    persisted = []

    def fake_persist(accounts, **kwargs):
        persisted.extend(accounts)
        return {
            "status": "completed",
            "upserted": len(accounts),
            "database": "storage-intelligence",
            "container": "storage-accounts",
        }

    monkeypatch.setattr(web_app, "persist_inventory_accounts", fake_persist)
    try:
        response = TestClient(web_app.app).post(
            "/api/accounts/import",
            files={"spreadsheet": ("airgap.csv", content, "text/csv")},
        )
        assert response.status_code == 201
        assert persisted[0]["account_id"] == account_id
        assert persisted[0]["subscription_id"] == subscription_id
    finally:
        web_app.ACCOUNTS[:] = [row for row in web_app.ACCOUNTS if row["name"] != account_name]


def test_bulk_import_rejects_mismatched_account_subscription_ids(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    import web.app as web_app

    account_name = "stairgapmismatch01"
    account_id = (
        "/subscriptions/22222222-2222-2222-2222-222222222222/resourceGroups/rg-airgap/"
        f"providers/Microsoft.Storage/storageAccounts/{account_name}"
    )
    content = (
        "account_id,name,tenant_id,subscription_id,management_group,subscription,"
        "environment,subsidiary,region,tier\n"
        f"{account_id},{account_name},{TENANTS[0]['id']},"
        f"33333333-3333-3333-3333-333333333333,{SUBSCRIPTIONS[0]['management_group']},"
        f"platform-prod,Prod,{SUBSCRIPTIONS[0]['subsidiary']},eastus2,Hot\n"
    ).encode()
    response = TestClient(web_app.app).post(
        "/api/accounts/import",
        files={"spreadsheet": ("airgap.csv", content, "text/csv")},
    )

    assert response.status_code == 422
    assert "subscription_id does not match" in str(response.json()["detail"])
    assert not any(row["name"] == account_name for row in web_app.ACCOUNTS)


def test_catalog_management_and_all_azure_regions(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import (
        ACCOUNTS,
        BUSINESS_UNIT_CATALOG,
        MANAGEMENT_GROUP_CATALOG,
        REGION_CATALOG,
        SUBSCRIPTION_CATALOG,
        TENANT_CATALOG,
        app,
    )

    client = TestClient(app)
    portfolio = client.get("/api/portfolio").json()
    assert len(portfolio["filters"]["regions"]) >= 50
    assert portfolio["catalog"]["region_labels"]["swedencentral"] == "Sweden Central"
    assert portfolio["catalog"]["region_labels"]["chilecentral"] == "Chile Central"

    try:
        subscription = client.post("/api/catalog/subscriptions", json={"value": "beverage-prod"})
        business_unit = client.post("/api/catalog/business-units", json={"value": "Beverages"})
        management_group = client.post("/api/catalog/management-groups", json={"value": "mg-beverages"})
        tenant = client.post("/api/catalog/tenants", json={"value": "44444444-4444-4444-8444-444444444444"})
        region = client.post("/api/catalog/regions", json={"value": "Chile Central"})
        assert subscription.status_code == 201
        assert business_unit.status_code == 201
        assert management_group.status_code == 201
        assert tenant.status_code == 201
        assert region.status_code == 201
        assert region.json()["value"] == "chilecentral"
        assert client.post("/api/catalog/subscriptions", json={"value": "beverage-prod"}).status_code == 409

        account = client.post(
            "/api/accounts",
            json={
                "name": "stbeveragecl01",
                "tenant_id": "44444444-4444-4444-8444-444444444444",
                "management_group": "mg-beverages",
                "subscription": "beverage-prod",
                "environment": "Prod",
                "subsidiary": "Beverages",
                "region": "chilecentral",
                "tier": "Cool",
            },
        )
        assert account.status_code == 201
        assert account.json()["account"]["region"] == "chilecentral"
    finally:
        ACCOUNTS[:] = [row for row in ACCOUNTS if row["name"] != "stbeveragecl01"]
        SUBSCRIPTION_CATALOG.discard("beverage-prod")
        BUSINESS_UNIT_CATALOG.discard("Beverages")
        MANAGEMENT_GROUP_CATALOG.discard("mg-beverages")
        TENANT_CATALOG.discard("44444444-4444-4444-8444-444444444444")
        REGION_CATALOG.discard("chilecentral")


def test_azure_cli_tenant_discovery(monkeypatch):
    from storage_intelligence import discovery

    monkeypatch.delenv("AZURE_CLIENT_ID", raising=False)

    def fake_az(arguments, timeout=180):
        del timeout
        if arguments[:3] == ["account", "list", "--all"]:
            return [
                {
                    "id": "11111111-1111-1111-1111-111111111111",
                    "name": "tenant-platform",
                    "tenantId": "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
                    "managementGroup": "mg-platform",
                    "state": "Enabled",
                }
            ]
        if arguments[:3] == ["account", "management-group", "list"]:
            return [{"name": "mg-platform"}]
        if arguments[:4] == ["account", "management-group", "subscription", "show-sub-under-mg"]:
            return [{"name": "11111111-1111-1111-1111-111111111111"}]
        if arguments[:3] == ["storage", "account", "list"]:
            return [
                {
                    "id": "/subscriptions/11111111-1111-1111-1111-111111111111/resourceGroups/rg/providers/Microsoft.Storage/storageAccounts/stdiscovered01",
                    "name": "stdiscovered01",
                    "primaryLocation": "swedencentral",
                    "accessTier": "Cool",
                    "allowSharedKeyAccess": True,
                    "publicNetworkAccess": "Enabled",
                    "allowBlobPublicAccess": True,
                    "isHnsEnabled": True,
                    "isSftpEnabled": True,
                    "privateEndpointConnections": [],
                    "kind": "StorageV2",
                    "sku": {"name": "Standard_ZRS"},
                    "tags": {
                        "BusinessUnit": "Beverages",
                        "ManagementGroup": "mg-beverages",
                        "Environment": "QA",
                        "DatabricksWorkspace": "dbw-beverages",
                        "FabricLakehouse": "lakehouse-beverages",
                        "SAPSystem": "SAP-S4HANA-BEVERAGES",
                        "AzureDataFactory": "adf-beverages-prod",
                        "UsesSASKeys": "true",
                        "ServicePrincipalAccess": "false",
                        "NetworkSecurityGroup": "nsg-beverages",
                        "ApplicationSecurityGroup": "asg-beverages",
                        "Project": "beverages-modernization",
                        "LastAccessedDate": "2024-02-01",
                        "ProjectStatus": "defunct",
                        "ApplicationInsights": "appi-beverages-prod",
                    },
                }
            ]
        raise AssertionError(arguments)

    monkeypatch.setattr(discovery, "_run_az", fake_az)
    result = discovery.discover_storage_accounts(["aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"])
    assert result["subscriptions"] == 1
    assert result["accounts"][0]["name"] == "stdiscovered01"
    assert result["accounts"][0]["business_unit"] == "Beverages"
    assert result["accounts"][0]["subsidiary"] == "Beverages"
    assert result["accounts"][0]["management_group"] == "mg-beverages"
    assert result["accounts"][0]["environment"] == "QA"
    assert result["management_groups"] == ["mg-beverages"]
    assert result["subsidiaries"] == ["Beverages"]
    assert result["environments"] == ["QA"]
    assert result["accounts"][0]["region"] == "swedencentral"
    assert result["accounts"][0]["tier"] == "Cool"
    assert result["accounts"][0]["databricks_workspace"] == "dbw-beverages"
    assert result["accounts"][0]["fabric_lakehouse"] == "lakehouse-beverages"
    assert result["accounts"][0]["sap_system"] == "SAP-S4HANA-BEVERAGES"
    assert result["accounts"][0]["azure_data_factory"] == "adf-beverages-prod"
    assert result["accounts"][0]["uses_sas_keys"] is True
    assert result["accounts"][0]["shared_key_access_enabled"] is True
    assert result["accounts"][0]["public_network_access"] is True
    assert result["accounts"][0]["blob_public_access_enabled"] is True
    assert result["accounts"][0]["private_endpoint_enabled"] is False
    assert result["accounts"][0]["service_principal_access_enabled"] is False
    assert result["accounts"][0]["network_security_group"] == "nsg-beverages"
    assert result["accounts"][0]["application_security_group"] == "asg-beverages"
    assert result["accounts"][0]["project_name"] == "beverages-modernization"
    assert result["accounts"][0]["tag_business_unit"] == "Beverages"
    assert result["accounts"][0]["last_accessed_date"] == "2024-02-01"
    assert result["accounts"][0]["project_defunct"] is True
    assert result["accounts"][0]["hns_enabled"] is True
    assert result["accounts"][0]["sftp_enabled"] is True
    assert result["accounts"][0]["application_insights_resource"] == "appi-beverages-prod"


def test_admin_tenant_discovery_ingests_accounts(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web import app as web_app

    discovered_name = "stadminpull01"
    monkeypatch.setattr(
        web_app,
        "discover_storage_accounts",
        lambda: {
            "pulled_at": "2026-08-11T20:00:00+00:00",
            "tenants": ["aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"],
            "subscriptions": 1,
            "accounts": [
                {
                    "account_id": "/subscriptions/sub-id/resourceGroups/rg/providers/Microsoft.Storage/storageAccounts/stadminpull01",
                    "name": discovered_name,
                    "tenant_id": "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
                    "subscription_id": "sub-id",
                    "subscription_name": "admin-discovered-subscription",
                    "environment": "Perf",
                    "management_group": "mg-admin",
                    "subsidiary": "Beverages",
                    "business_unit": "Beverages",
                    "region": "swedencentral",
                    "tier": "Cool",
                    "tier_assumed": False,
                    "kind": "StorageV2",
                    "sku": "Standard_ZRS",
                    "databricks_workspace": "dbw-admin",
                    "fabric_lakehouse": "lakehouse-admin",
                }
            ],
        },
    )
    persisted = []

    def fake_persist(accounts, *, pulled_at, trigger):
        persisted.extend(accounts)
        assert pulled_at == "2026-08-11T20:00:00+00:00"
        assert trigger == "manual"
        return {
            "status": "completed",
            "upserted": len(accounts),
            "database": "storage-intelligence",
            "container": "storage-accounts",
        }

    monkeypatch.setattr(web_app, "persist_inventory_accounts", fake_persist)
    try:
        response = TestClient(web_app.app).post("/api/admin/discovery/pull")
        assert response.status_code == 202
        status = TestClient(web_app.app).get("/api/admin/discovery/status")
        assert status.json()["status"] == "completed"
        assert status.json()["added"] == 1
        assert status.json()["persisted"] == 1
        assert status.json()["environments"] == 1
        assert persisted[0]["name"] == discovered_name
        assert any(row["name"] == discovered_name for row in web_app.ACCOUNTS)
    finally:
        web_app.ACCOUNTS[:] = [row for row in web_app.ACCOUNTS if row["name"] != discovered_name]
        web_app.SUBSCRIPTION_CATALOG.discard("admin-discovered-subscription")
        web_app.BUSINESS_UNIT_CATALOG.discard("Beverages")
        web_app.MANAGEMENT_GROUP_CATALOG.discard("mg-admin")
        web_app.TENANT_CATALOG.discard("aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa")


def test_cosmos_inventory_upserts_with_managed_identity(monkeypatch):
    from storage_intelligence.cosmos_inventory import persist_inventory_accounts

    monkeypatch.setenv("COSMOS_INVENTORY_ENABLED", "true")
    monkeypatch.setenv("COSMOS_ENDPOINT", "https://cosmos.example/")
    monkeypatch.setenv("COSMOS_DATABASE", "storage-intelligence")
    monkeypatch.setenv("COSMOS_CONTAINER", "storage-accounts")
    monkeypatch.setenv("AZURE_CLIENT_ID", "managed-identity-client-id")
    upserted = []
    credentials = []

    class FakeContainer:
        def upsert_item(self, *, body):
            upserted.append(body)

    class FakeDatabase:
        def get_container_client(self, name):
            assert name == "storage-accounts"
            return FakeContainer()

    class FakeClient:
        def get_database_client(self, name):
            assert name == "storage-intelligence"
            return FakeDatabase()

    def client_factory(endpoint, *, credential):
        assert endpoint == "https://cosmos.example/"
        assert credential == "credential"
        return FakeClient()

    def credential_factory(**kwargs):
        credentials.append(kwargs)
        return "credential"

    result = persist_inventory_accounts(
        [
            {
                "account_id": "/subscriptions/sub-id/resourceGroups/rg/providers/Microsoft.Storage/storageAccounts/stpersisted01",
                "name": "stpersisted01",
                "subscription_id": "sub-id",
                "environment": "Dev",
                "tenant_id": TENANTS[0]["id"],
                "management_group": SUBSCRIPTIONS[0]["management_group"],
                "subsidiary": SUBSCRIPTIONS[0]["subsidiary"],
                "region": "swedencentral",
                "tier": "Cool",
                "uses_sas_keys": True,
                "public_network_access": True,
                "private_endpoint_enabled": False,
                "project_name": "cosmos-project",
                "project_defunct": False,
                "hns_enabled": True,
                "sftp_enabled": True,
                "application_insights_resource": "appi-cosmos",
            }
        ],
        pulled_at="2026-08-12T19:00:00+00:00",
        trigger="schedule",
        client_factory=client_factory,
        credential_factory=credential_factory,
    )

    assert result["status"] == "completed"
    assert result["upserted"] == 1
    assert credentials == [{"managed_identity_client_id": "managed-identity-client-id"}]
    assert upserted[0]["subscription_id"] == "sub-id"
    assert upserted[0]["environment"] == "Dev"
    assert upserted[0]["tenant_id"] == TENANTS[0]["id"]
    assert upserted[0]["management_group"] == SUBSCRIPTIONS[0]["management_group"]
    assert upserted[0]["subsidiary"] == SUBSCRIPTIONS[0]["subsidiary"]
    assert upserted[0]["uses_sas_keys"] is True
    assert upserted[0]["private_endpoint_enabled"] is False
    assert upserted[0]["project_name"] == "cosmos-project"
    assert upserted[0]["sftp_enabled"] is True
    assert upserted[0]["application_insights_resource"] == "appi-cosmos"
    assert upserted[0]["source"] == "azure-cli-discovery-v1"
    assert upserted[0]["discovery_trigger"] == "schedule"
    assert upserted[0]["resource_id"].endswith("/stpersisted01")
    assert "/" not in upserted[0]["id"]


def test_cosmos_inventory_requires_endpoint_when_enabled(monkeypatch):
    from storage_intelligence.cosmos_inventory import persist_inventory_accounts

    monkeypatch.setenv("COSMOS_INVENTORY_ENABLED", "true")
    monkeypatch.delenv("COSMOS_ENDPOINT", raising=False)

    with pytest.raises(RuntimeError, match="COSMOS_ENDPOINT is required"):
        persist_inventory_accounts([], pulled_at="2026-08-12T19:00:00+00:00", trigger="schedule")


def test_admin_tenant_discovery_requires_role(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "false")
    from web.app import app

    principal = base64.b64encode(
        json.dumps({"identityProvider": "aad", "claims": []}).encode()
    ).decode()
    response = TestClient(app).post(
        "/api/admin/discovery/pull",
        headers={"x-ms-client-principal": principal},
    )
    assert response.status_code == 403


def test_admin_can_configure_discovery_cron(monkeypatch, tmp_path):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web import app as web_app

    original_schedule = web_app.DISCOVERY_SCHEDULE["cron"]
    monkeypatch.setattr(
        web_app,
        "DISCOVERY_SCHEDULE_PATH",
        tmp_path / "discovery-schedule.json",
    )
    try:
        response = TestClient(web_app.app).put(
            "/api/admin/discovery/schedule",
            json={"cron": "15 */2 * * *"},
        )
        assert response.status_code == 200
        assert response.json()["schedule"] == "15 */2 * * *"
        assert response.json()["next_run"]
        assert json.loads(web_app.DISCOVERY_SCHEDULE_PATH.read_text())["cron"] == "15 */2 * * *"
        invalid = TestClient(web_app.app).put(
            "/api/admin/discovery/schedule",
            json={"cron": "not a cron expression"},
        )
        assert invalid.status_code == 422
    finally:
        web_app.DISCOVERY_SCHEDULE["cron"] = original_schedule
        web_app.DISCOVERY_STATE["schedule"] = original_schedule


def test_admin_discovery_has_role_gated_navigation():
    app_script = (Path(__file__).parents[1] / "src" / "web" / "static" / "app.js").read_text(encoding="utf-8")
    styles = (Path(__file__).parents[1] / "src" / "web" / "static" / "styles.css").read_text(encoding="utf-8")
    assert '{ id: "admin", label: "Admin"' in app_script
    assert 'view.id !== "admin" || (portfolio && portfolio.permissions.admin)' in app_script
    assert 'activeView === "admin" && portfolio.permissions.admin' in app_script
    assert "Tenant-wide storage account discovery & Retrieval" in app_script
    assert "Retrieve Storage Inventory" in app_script
    assert "Pull Tenant Wide Storage Account Details" not in app_script
    assert 'className: "schedule-examples", role: "table"' in app_script
    assert 'className: "schedule-example-row", role: "row"' in app_script
    assert 'className: "data-row admin-schedule-row"' not in app_script
    assert ".schedule-example-header, .schedule-example-row {" in styles
    assert ".schedule-example-row { grid-template-columns: 1fr; gap: 5px; }" in styles
    assert "function accountClassification(account)" in app_script
    assert "!hasPlatformClassification(account) && e(AccountClassificationLogo" in app_script
    assert 'return { key: "data", label: "Azure data and analytics workload" };' in app_script
    assert ".classification-data .classification-glyph" in styles
    assert 'className: "savings-table-header", role: "row"' in app_script
    assert '"Current tier"' in app_script
    assert '"Recommended target tier"' in app_script
    assert "account.current_size_tb" in app_script
    assert ".data-row.savings-row" in styles


def test_airgap_sample_workbook_has_complete_header_and_imports(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import ACCOUNTS, app

    expected_headers = [
        "account_id",
        "name",
        "tenant_id",
        "subscription_id",
        "subscription_name",
        "environment",
        "management_group",
        "subsidiary",
        "business_unit",
        "region",
        "tier",
        "tier_assumed",
        "kind",
        "sku",
        "uses_sas_keys",
        "shared_key_access_enabled",
        "public_network_access",
        "blob_public_access_enabled",
        "private_endpoint_enabled",
        "service_principal_access_enabled",
        "managed_identity_enabled",
        "network_security_group",
        "application_security_group",
        "project_name",
        "tag_business_unit",
        "last_accessed_date",
        "project_defunct",
        "databricks_workspace",
        "fabric_lakehouse",
        "sap_system",
        "azure_data_factory",
        "hns_enabled",
        "sftp_enabled",
        "application_insights_resource",
        "azure_function_app",
        "log_analytics_workspace",
    ]
    sample_path = Path(__file__).parents[1] / "src" / "web" / "static" / "Sample.xlsx"
    workbook = load_workbook(sample_path, read_only=True, data_only=True)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()

    assert rows == [tuple(expected_headers)]
    response = TestClient(app).get("/static/Sample.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.content.startswith(b"PK")

    import_name = "stairgapsample01"
    workbook = load_workbook(sample_path)
    workbook.active.append(
        [
            "",
            import_name,
            TENANTS[0]["id"],
            "",
            "platform-prod",
            "Prod",
            SUBSCRIPTIONS[0]["management_group"],
            SUBSCRIPTIONS[0]["subsidiary"],
            SUBSCRIPTIONS[0]["subsidiary"],
            "eastus2",
            "Hot",
            False,
            "StorageV2",
            "Standard_ZRS",
        ]
    )
    payload = BytesIO()
    workbook.save(payload)
    workbook.close()
    try:
        imported = TestClient(app).post(
            "/api/accounts/import",
            files={
                "spreadsheet": (
                    "Sample.xlsx",
                    payload.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert imported.status_code == 201
        assert imported.json()["accounts"] == [import_name]
    finally:
        ACCOUNTS[:] = [row for row in ACCOUNTS if row["name"] != import_name]


def test_hierarchy_controls_and_foundry_mark_are_global():
    static_root = Path(__file__).parents[1] / "src" / "web" / "static"
    app_script = (static_root / "app.js").read_text(encoding="utf-8")
    styles = (static_root / "styles.css").read_text(encoding="utf-8")
    translations = (static_root / "translations.js").read_text(encoding="utf-8")
    index_html = (static_root / "index.html").read_text(encoding="utf-8")
    foundry_logo = static_root / "assets" / "azure-ai-foundry.png"
    navigation_icons = {
        "overview": "nav-overview.svg",
        "agent": "nav-agent.svg",
        "savings": "nav-savings.svg",
        "findings": "nav-findings.svg",
        "health": "nav-health.svg",
        "admin": "nav-admin.svg",
    }
    assert 'function FoundryMiniLogo()' in app_script
    assert "function EntraAuthenticatedMark()" in app_script
    assert 'src: "/static/assets/microsoft-entra-id.png"' in app_script
    assert "width: 22px; height: 22px;" in styles
    assert ".entra-id-logo { width: 14px; height: 14px;" in styles
    assert 'className: "entra-lock"' not in app_script
    assert 'e("div", { className: "identity" }' not in app_script
    entra_logo = static_root / "assets" / "microsoft-entra-id.png"
    assert entra_logo.exists()
    assert hashlib.sha256(entra_logo.read_bytes()).hexdigest().upper() == (
        "407C905DE2C610918819D3E5138664E902F4F6D097238EC74EDA62F3D7665767"
    )
    assert 'src: "/static/assets/azure-ai-foundry.png"' in app_script
    assert 'alt: "Azure AI Foundry"' in app_script
    assert foundry_logo.exists()
    assert hashlib.sha256(foundry_logo.read_bytes()).hexdigest().upper() == (
        "E2C39AF95049197F78F9FE1057D99C625E3E997EEBA34754DFC28D771731605D"
    )
    assert 'label: "tenant IDs"' in app_script
    assert 'label: "management groups"' in app_script
    assert 'label: "environments"' in app_script
    assert 'label: "subsidiaries / business units"' in app_script
    assert 'className: "hierarchy-context"' in app_script
    assert "function SapLogo()" in app_script
    assert "function DataFactoryLogo()" in app_script
    assert "function SftpLogo()" in app_script
    assert "function AppInsightsLogo()" in app_script
    assert "function FunctionAppLogo()" in app_script
    assert "function LogAnalyticsLogo()" in app_script
    assert "function-app-link" in app_script
    assert "log-analytics-link" in app_script
    assert "Managed identity enabled" in app_script
    assert "Managed identity disabled" in app_script
    assert "function NavIcon(props)" in app_script
    assert "select { color: #8292a3; }" in styles
    assert "select:focus, select:active { color: var(--text); }" in styles
    assert "select option { color: var(--text); background: var(--panel-2); }" in styles
    for view_id, filename in navigation_icons.items():
        path = static_root / "assets" / filename
        assert path.exists()
        assert path.read_text(encoding="utf-8").startswith("<svg")
        assert f'id: "{view_id}"' in app_script
        assert f'icon: "/static/assets/{filename}"' in app_script
    view_order = ["overview", "health", "findings", "savings", "agent", "admin"]
    assert [app_script.index(f'id: "{view_id}"') for view_id in view_order] == sorted(
        app_script.index(f'id: "{view_id}"') for view_id in view_order
    )
    assert 'subtitle: "Capacity, Cost, Risk, Impact, And Defensible Savings."' in app_script
    assert 'subtitle: "Review Prioritized Risks, Anomalies, Freshness Issues, And Savings Actions."' in app_script
    assert 'subtitle: "Inspect Source Status, Freshness Coverage, Stale Accounts, And Quality Gaps."' in app_script
    assert 'e("div", { className: "authored" }, "Authored by nrp")' in app_script
    assert 'className: "protocol-enabled"' in app_script
    assert '"aria-label": "MCP and A2A protocols enabled"' in app_script
    assert '}, "MCP & A2A Enabled")' in app_script
    assert '"MCP & A2A Enabled": "MCP y A2A habilitados"' in translations
    assert '{ key: "security", label: "Security"' in app_script
    assert '{ key: "governance", label: "Governance"' in app_script
    assert 'fetch("/api/questions")' in app_script
    assert "Save question" in app_script
    assert "function PostureMetric(props)" in app_script
    assert 'fetch("/api/posture/" + postureSelection' in app_script
    assert 'className: "posture-account-list risk-scroll"' in app_script
    assert 'className: "platform-account-grid platform-account-scroll"' in app_script
    assert '"aria-label": "Platform-linked storage accounts"' in app_script
    assert "portfolio.platform_accounts.length.toLocaleString()" in app_script
    assert '"Import account spreadsheet"' in app_script
    assert 'e("em", { className: "airgap-import-qualifier" }, "(for AIRGAP Accounts ONLY, if applicable)")' in app_script
    assert "Upload account spreadsheet (AIRGAP Accounts if any)" not in app_script
    assert 'href: "/static/Sample.xlsx"' in app_script
    assert 'download: "Sample.xlsx"' in app_script
    assert 'className: "sample-spreadsheet-link"' in app_script
    assert "Use XLSX or UTF-8 CSV with columns:" not in app_script
    assert '"aria-label": "AIRGAP account spreadsheet"' in app_script
    assert '"Import account spreadsheet": "Importar hoja de cálculo de cuentas"' in translations
    assert (
        '"(for AIRGAP Accounts ONLY, if applicable)": '
        '"(SOLO para cuentas AIRGAP, si corresponde)"'
    ) in translations
    assert '"Upload account spreadsheet (AIRGAP Accounts if any)":' not in translations
    assert "Use XLSX or UTF-8 CSV with columns:" not in translations
    assert "max-height: 560px; overflow-y: scroll;" in styles
    assert ".platform-account-scroll::-webkit-scrollbar-thumb" in styles
    assert 'className: "panel risk-overview-panel"' in app_script
    assert '"Risk concentration and account findings"' in app_script
    assert 'e("details", { className: "risk-account-row"' in app_script
    assert 'className: "risk-account-details"' in app_script
    assert 'className: "risk-component-grid"' in app_script
    assert 'className: "risk-factor-list"' in app_script
    assert "row.risk_factors.map" in app_script
    assert '"Priority findings"' not in app_script
    assert '"Priority findings":' not in translations
    assert "priority-findings" not in styles
    assert ".risk-pie-legend { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr));" in styles
    assert ".risk-component-grid { display: grid; grid-template-columns: repeat(4, minmax(110px, 1fr));" in styles
    assert "max-height: 640px; overflow-y: auto;" in styles
    assert "display: grid; place-items: center; min-height: 54px; font-size: 12px;" in styles
    for label in (
        "SAS Key",
        "Public Access",
        "No Private Endpoint",
        "No Service Principal",
        "Managed Identity",
        "No GRS/GZRS",
        "NSG/ASG linked",
        "Defunct Projects",
        "SFTP Enabled",
        "AppInsights Data",
        "Stale accounts",
        "Missing lifecycle",
    ):
        assert f'label: "{label}"' in app_script
    overview_segment = app_script[
        app_script.index('activeView === "overview"'):
        app_script.index('activeView === "admin"')
    ]
    health_segment = app_script[app_script.index('activeView === "health"'):]
    assert "e(PostureMetric" not in overview_segment
    assert "e(PostureDrilldown" not in overview_segment
    assert "e(PostureDrilldown" in health_segment
    assert health_segment.count("e(PostureMetric") == 12
    assert "Stale account details" not in health_segment
    assert health_segment.index('className: "source-grid"') < health_segment.index(
        'className: "metrics health-metrics"'
    )
    assert index_html.index("/static/translations.js?v=20260828-powered-by") < index_html.index(
        "/static/app.js?v=20260831-agent-usage"
    )
    assert "/static/styles.css?v=20260831-agent-usage" in index_html
    assert "/static/app.js?v=20260831-agent-usage" in index_html
    assert "<title>Storage Atlas</title>" in index_html
    assert 'title: "Overview", subtitle:' in app_script
    assert 'e("div", { className: "product-name" }, "Storage Atlas")' in app_script
    assert 'className: "avidunixuser-wordmark"' not in app_script
    assert ".brand { display: flex; flex-direction: column; align-items: center;" in styles
    assert ".avidunixuser-logo { width: 64px; height: 64px;" in styles
    assert ".product-name { color: #173b67; font-size: 21px;" in styles
    assert ".mobile-brand { display: none; flex-direction: column; align-items: center;" in styles
    assert "h1 { margin: 7px 0 5px; font-size: clamp(20px, 2.5vw, 28px); font-weight: 500;" in styles
    assert 'e("span", { className: "powered-label" }, "Powered by")' in app_script
    assert ".powered-label { color: rgba(80,101,122,.72); font-size: 9px;" in styles
    assert 'function AgentUsageTile(props)' in app_script
    assert 'fetch("/api/agent/query"' in app_script
    assert 'e(AgentUsageTile, { agent: agentUsage })' in app_script
    assert 'e("span", { className: "agent-usage-label" }, "Model:")' in app_script
    assert 'e("span", { className: "agent-usage-label" }, "Tokens:")' in app_script
    assert 'e("span", { className: "agent-usage-label" }, "Context used:")' in app_script
    assert ".agent-usage-tile {" in styles
    assert '"Powered by": "Con tecnología de"' in translations
    assert '"Storage Atlas": "Storage Atlas"' in translations
    assert 'function AvidunixuserLogo()' in app_script
    assert 'src: "/static/assets/avidunixuser-logo.png?v=20260901-light"' in app_script
    assert 'alt: "Avidunixuser"' in app_script
    avidunixuser_logo = static_root / "assets" / "avidunixuser-logo.png"
    assert avidunixuser_logo.exists()
    assert hashlib.sha256(avidunixuser_logo.read_bytes()).hexdigest().upper() == (
        "EF8011511CD516A0B9A9E67B22D4A35BDE7CCC5662570913A8FC8CCCE20377B6"
    )
    assert '<meta name="theme-color" content="#0078d4">' in index_html
    assert (
        'font-family: "Segoe UI Variable", "Segoe UI", Inter, ui-sans-serif, system-ui, '
        '-apple-system, BlinkMacSystemFont, sans-serif;'
    ) in styles
    assert "--accent: #0078d4;" in styles
    assert "--accent-hover: #106ebe;" in styles
    assert "--accent-pressed: #005a9e;" in styles
    assert "background: linear-gradient(135deg, var(--cyan), var(--green));" not in styles
    assert ".ask:hover:not(:disabled)" in styles
    assert ".notify-owners-button:hover:not(:disabled)" in styles
    assert 'localStorage.setItem("storage-intelligence-language", language)' in app_script
    assert "document.documentElement.lang = language" in app_script
    assert 'className: "language-switch"' in app_script
    assert 'e("option", { value: "en" }, "English")' in app_script
    assert 'e("option", { value: "es" }, "Spanish")' in app_script
    for english, spanish in (
        ("Overview", "Resumen"),
        ("Agent Investigation", "Investigación del agente"),
        ("Savings Simulator", "Simulador de ahorros"),
        ("Findings", "Hallazgos"),
        ("Data Health", "Salud de los datos"),
        ("Admin", "Administración"),
        ("Capacity, Cost, Risk, Impact, And Defensible Savings.", "Capacidad, Costo, Riesgo, Impacto Y Ahorros Justificables."),
        ("Review Prioritized Risks, Anomalies, Freshness Issues, And Savings Actions.", "Revise Riesgos Priorizados, Anomalías, Problemas De Vigencia Y Acciones De Ahorro."),
        ("Inspect Source Status, Freshness Coverage, Stale Accounts, And Quality Gaps.", "Inspeccione El Estado De Las Fuentes, La Cobertura De Vigencia, Las Cuentas Obsoletas Y Las Brechas De Calidad."),
        ("Total Findings", "Total de hallazgos"),
        ("Stale accounts", "Cuentas obsoletas"),
        ("Save question", "Guardar pregunta"),
        ("Retrieve Storage Inventory", "Recuperar inventario de almacenamiento"),
        ("Recommended target tier", "Nivel objetivo recomendado"),
    ):
        assert f'"{english}": "{spanish}"' in translations
    assert "function FindingMetric(props)" in app_script
    assert 'className: "finding-filter-bar"' not in app_script
    for label in (
        "Total Findings",
        "Data Freshness",
        "Growth Anomaly",
        "Risk",
        "Savings Action",
    ):
        assert f'label: "{label}"' in app_script


def test_saved_question_api_persists_locally(monkeypatch, tmp_path):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    monkeypatch.setenv("COSMOS_INVENTORY_ENABLED", "false")
    monkeypatch.setenv("SAVED_QUESTIONS_PATH", str(tmp_path / "saved-questions.json"))
    from storage_intelligence.question_library import DEFAULT_QUESTIONS
    from web.app import app

    client = TestClient(app)
    initial = client.get("/api/questions")
    assert initial.status_code == 200
    assert initial.json()["defaults"] == len(DEFAULT_QUESTIONS)
    assert initial.json()["custom"] == 0

    question = "Which storage accounts have cost spikes after unusual transaction bursts?"
    saved = client.post("/api/questions", json={"question": question})
    assert saved.status_code == 201
    assert saved.json()["saved"]["question"] == question
    assert saved.json()["custom"] == 1

    reloaded = client.get("/api/questions").json()
    assert reloaded["questions"][-1]["question"] == question
    assert reloaded["questions"][-1]["custom"] is True
    assert client.post("/api/questions", json={"question": f"  {question}  "}).status_code == 409


def test_saved_question_uses_cosmos_partition(monkeypatch):
    from storage_intelligence import question_library

    monkeypatch.setenv("COSMOS_INVENTORY_ENABLED", "true")
    records = []

    class FakeContainer:
        def query_items(self, *, query, parameters, partition_key):
            assert "saved-agent-question" not in query
            assert parameters[0]["value"] == "saved-agent-question"
            assert partition_key == question_library.QUESTION_PARTITION
            return list(records)

        def upsert_item(self, *, body):
            records.append(body)

    monkeypatch.setattr(question_library, "get_cosmos_container", lambda: FakeContainer())
    question = "Which accounts show simultaneous capacity growth and transaction pressure?"
    saved = question_library.save_question(question, "test-user")

    assert saved["question"] == question
    assert records[0]["subscription_id"] == question_library.QUESTION_PARTITION
    assert records[0]["document_type"] == "saved-agent-question"
    assert question_library.list_questions()[-1]["question"] == question


def test_functional_view_endpoints(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import app

    client = TestClient(app)
    savings = client.post(
        "/api/savings/simulate",
        json={"adoption_pct": 37, "filters": {"subsidiary": SUBSCRIPTIONS[0]["subsidiary"]}},
    )
    assert savings.status_code == 200
    assert savings.json()["simulation"]["adoption_pct"] == 37
    recommendation = savings.json()["simulation"]["top_accounts"][0]
    assert recommendation["current_tier"] in {"Hot", "Cool", "Cold", "Archive"}
    assert recommendation["target_tier"] in {"Cold", "Archive"}
    assert recommendation["current_size_tb"] > 0
    assert len(savings.json()["scenarios"]) == 3
    assert savings.json()["confidence"]["level"] in {"low", "medium", "high"}
    assert savings.json()["scope"]["filters"]["subsidiary"] == SUBSCRIPTIONS[0]["subsidiary"]
    assert savings.json()["scope"]["tenant_count"] == 1
    assert savings.json()["scope"]["management_group_count"] == 2
    assert savings.json()["scope"]["environment_count"] >= 1

    findings = client.get(
        "/api/findings",
        params={
            "tenant_id": TENANTS[0]["id"],
            "management_group": SUBSCRIPTIONS[0]["management_group"],
            "subsidiary": SUBSCRIPTIONS[0]["subsidiary"],
            "environment": "Prod",
        },
    )
    assert findings.status_code == 200
    assert findings.json()["scope"]["tenant_count"] == 1
    assert findings.json()["scope"]["management_group_count"] == 1
    assert findings.json()["scope"]["subsidiary_count"] == 1
    assert findings.json()["scope"]["environment_count"] == 1
    assert findings.json()["total"] == len(findings.json()["findings"])
    assert {"Risk", "Growth anomaly", "Data freshness", "Savings action"}.issubset(
        findings.json()["counts"]
    )
    assert set(findings.json()["counts"]) == {
        "Risk",
        "Growth anomaly",
        "Data freshness",
        "Savings action",
    }
    risk_finding = next(item for item in findings.json()["findings"] if item["category"] == "Risk")
    assert risk_finding["risk_factors"]
    assert "security" in risk_finding["summary"].lower()
    severities = {"high": 0, "medium": 1, "low": 2}
    assert findings.json()["findings"] == sorted(
        findings.json()["findings"],
        key=lambda item: (severities[item["severity"]], -item["value"]),
    )

    health = client.get(
        "/api/data-health",
        params={"subsidiary": SUBSCRIPTIONS[0]["subsidiary"]},
    )
    assert health.status_code == 200
    assert health.json()["scope"]["subsidiary_count"] == 1
    assert health.json()["summary"]["sas_key_accounts"] > 0
    assert health.json()["summary"]["public_access_accounts"] > 0
    assert health.json()["summary"]["missing_private_endpoint_accounts"] > 0
    assert health.json()["summary"]["missing_service_principal_access_accounts"] > 0
    assert health.json()["summary"]["managed_identity_accounts"] > 0
    assert health.json()["summary"]["non_geo_redundant_accounts"] > 0
    assert health.json()["summary"]["defunct_project_accounts"] > 0
    assert health.json()["summary"]["sftp_enabled_accounts"] > 0
    assert health.json()["summary"]["application_insights_accounts"] > 0
    assert 0 <= health.json()["summary"]["freshness_pct"] <= 100
    assert len(health.json()["sources"]) == 7
    assert any(source["name"] == "Azure Resource Graph" for source in health.json()["sources"])
    assert [source["name"] for source in health.json()["sources"]] == [
        "Synthetic pilot",
        "Azure CLI discovery",
        "Azure Resource Graph",
        "Blob Inventory",
        "Azure Monitor Metrics",
        "Cost Management exports",
        "Databricks system tables",
    ]


def test_admin_can_enable_and_run_connectors(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "true")
    from web.app import CONNECTOR_RUNTIME, app

    original = dict(CONNECTOR_RUNTIME["resource-graph"])
    client = TestClient(app)
    try:
        before = client.get("/api/data-health").json()
        resource_graph = next(
            source for source in before["sources"] if source["key"] == "resource-graph"
        )
        assert resource_graph["status"] == "disabled"
        assert resource_graph["records"] == 0
        assert resource_graph["eligible_records"] == 2500
        assert client.post("/api/admin/connectors/resource-graph/run").status_code == 409

        enabled = client.post("/api/admin/connectors/resource-graph/enable")
        assert enabled.status_code == 200
        assert enabled.json()["status"] == "enabled"
        run = client.post("/api/admin/connectors/resource-graph/run")
        assert run.status_code == 200
        assert run.json()["status"] == "healthy"
        assert run.json()["records"] == 2500

        after = client.get("/api/data-health").json()
        resource_graph = next(
            source for source in after["sources"] if source["key"] == "resource-graph"
        )
        assert resource_graph["status"] == "healthy"
        assert resource_graph["records"] == 2500
        assert resource_graph["last_run"]
    finally:
        CONNECTOR_RUNTIME["resource-graph"].clear()
        CONNECTOR_RUNTIME["resource-graph"].update(original)


def test_disabled_connector_status_is_the_enable_control():
    app_script = (Path(__file__).parents[1] / "src" / "web" / "static" / "app.js").read_text(encoding="utf-8")
    assert 'className: "source-status source-status-button disabled"' in app_script
    assert 'title: "Enable this connector so it can be run"' in app_script
    assert 'connectorAction === source.key ? "Enabling…" : "Disabled"' in app_script
    assert 'connectorAction === source.key ? "Running…" : "Run"' in app_script


def test_web_requires_entra_header(monkeypatch):
    monkeypatch.setenv("AUTH_DISABLED", "false")
    from web.app import app

    assert TestClient(app).get("/api/portfolio").status_code == 401


def test_foundry_agent_models_are_available():
    from azure.ai.projects.models import (
        OpenApiAgentTool,
        OpenApiFunctionDefinition,
        OpenApiManagedAuthDetails,
        OpenApiManagedSecurityScheme,
        PromptAgentDefinition,
    )

    assert all(
        (
            OpenApiAgentTool,
            OpenApiFunctionDefinition,
            OpenApiManagedAuthDetails,
            OpenApiManagedSecurityScheme,
            PromptAgentDefinition,
        )
    )


def test_preprovision_reuses_configured_entra_applications():
    scripts_root = Path(__file__).parents[1] / "scripts"
    powershell = (scripts_root / "preprovision.ps1").read_text(encoding="utf-8")
    shell = (scripts_root / "preprovision.sh").read_text(encoding="utf-8")

    assert '$env:WEB_AUTH_CLIENT_ID' in powershell
    assert '$env:FUNCTION_AUTH_CLIENT_ID' in powershell
    assert "--display-name $displayName" in powershell
    assert '"${WEB_AUTH_CLIENT_ID:-}"' in shell
    assert '"${FUNCTION_AUTH_CLIENT_ID:-}"' in shell
    assert '--display-name "$display_name"' in shell
