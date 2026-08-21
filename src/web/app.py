from __future__ import annotations

import base64
import asyncio
import csv
import json
import os
import re
import uuid
from contextlib import asynccontextmanager
from datetime import UTC, datetime
from io import BytesIO, StringIO
from pathlib import Path
from threading import Lock
from typing import Annotated, Any

from fastapi import BackgroundTasks, Depends, FastAPI, File, Header, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from croniter import croniter
from openpyxl import load_workbook
from pydantic import AliasChoices, BaseModel, ConfigDict, Field, ValidationError, field_validator

from storage_intelligence import IntelligenceEngine, generate_accounts
from storage_intelligence.analytics import (
    AT_RISK_THRESHOLD,
    freshness,
    growth_anomalies,
    portfolio_summary,
    risk_score,
    risks,
    savings_scenarios,
    tier_savings,
    top_actions,
)
from storage_intelligence.azure_regions import AZURE_REGION_LABELS
from storage_intelligence.cosmos_inventory import persist_discovered_accounts
from storage_intelligence.discovery import discover_storage_accounts
from storage_intelligence.hierarchy import (
    ENVIRONMENTS,
    MANAGEMENT_GROUPS,
    management_group_labels,
    subscription_by_name,
    tenant_labels,
)
from storage_intelligence.question_library import (
    QuestionAlreadyExistsError,
    QuestionLibraryFullError,
    list_questions,
    save_question,
)
from storage_intelligence.synthetic import dataset_fingerprint
from protocols.a2a_server import register_a2a_routes
from protocols.mcp_server import build_mcp_http_app, build_mcp_server
from protocols.service import StorageIntelligenceService
from web.notifications import (
    NotificationConfigurationError,
    NotificationDeliveryError,
    send_project_owner_notification,
)

STATIC = Path(__file__).parent / "static"
ACCOUNTS = generate_accounts()
ENGINE = IntelligenceEngine(ACCOUNTS)
AGENT_SERVICE = StorageIntelligenceService(ENGINE)
MCP_SERVER = build_mcp_server(ENGINE)
ACCOUNT_LOCK = Lock()
SUBSCRIPTION_CATALOG = {row["subscription"] for row in ACCOUNTS}
TENANT_CATALOG = {row["tenant_id"] for row in ACCOUNTS}
MANAGEMENT_GROUP_CATALOG = {row["management_group"] for row in ACCOUNTS}
SUBSIDIARY_CATALOG = {row["subsidiary"] for row in ACCOUNTS}
BUSINESS_UNIT_CATALOG = SUBSIDIARY_CATALOG
ENVIRONMENT_CATALOG = set(ENVIRONMENTS)
REGION_CATALOG = {row["region"] for row in ACCOUNTS}
TENANT_LABELS = tenant_labels()
MANAGEMENT_GROUP_LABELS = management_group_labels()
MANAGEMENT_GROUP_BY_ID = {group["id"]: group for group in MANAGEMENT_GROUPS}
SUBSCRIPTION_BY_NAME = subscription_by_name()
DISCOVERY_LOCK = Lock()
CONNECTOR_LOCK = Lock()
CONNECTOR_DEFINITIONS = {
    "resource-graph": {
        "name": "Azure Resource Graph",
        "flag": "ENABLE_RESOURCE_GRAPH",
        "detail": "Storage account resource and configuration inventory",
    },
    "blob-inventory": {
        "name": "Blob Inventory",
        "flag": "ENABLE_BLOB_INVENTORY",
        "detail": "Container, blob, tier, and retention inventory",
    },
    "azure-monitor": {
        "name": "Azure Monitor Metrics",
        "flag": "ENABLE_AZURE_MONITOR_METRICS",
        "detail": "Capacity, transactions, latency, and availability metrics",
    },
    "cost-exports": {
        "name": "Cost Management exports",
        "flag": "ENABLE_COST_EXPORTS",
        "detail": "Storage cost allocation and trend records",
    },
    "databricks": {
        "name": "Databricks system tables",
        "flag": "ENABLE_DATABRICKS_EXPORTS",
        "detail": "Workspace, external-location, IO, and small-file attribution",
    },
}
CONNECTOR_RUNTIME = {
    key: {
        "enabled": os.getenv(definition["flag"], "false").lower() == "true",
        "status": "enabled" if os.getenv(definition["flag"], "false").lower() == "true" else "disabled",
        "records": 0,
        "last_run": None,
        "mode": "pilot-fixture",
    }
    for key, definition in CONNECTOR_DEFINITIONS.items()
}
DEFAULT_DISCOVERY_CRON = "0 */6 * * *"
DISCOVERY_SCHEDULE_PATH = Path(
    os.getenv("DISCOVERY_SCHEDULE_PATH", "data/discovery-schedule.json")
)


def _load_discovery_schedule() -> str:
    configured = os.getenv("DISCOVERY_CRON", DEFAULT_DISCOVERY_CRON)
    if DISCOVERY_SCHEDULE_PATH.exists():
        try:
            saved = json.loads(DISCOVERY_SCHEDULE_PATH.read_text(encoding="utf-8"))
            configured = str(saved["cron"])
        except (OSError, KeyError, TypeError, json.JSONDecodeError):
            configured = os.getenv("DISCOVERY_CRON", DEFAULT_DISCOVERY_CRON)
    return configured if len(configured.split()) == 5 and croniter.is_valid(configured) else DEFAULT_DISCOVERY_CRON


DISCOVERY_SCHEDULE = {"cron": _load_discovery_schedule()}
DISCOVERY_STATE: dict[str, Any] = {
    "status": "idle",
    "schedule": DISCOVERY_SCHEDULE["cron"],
}
AGENT_STATE: dict[str, Any] = {"status": "not-requested"}
FUNCTION_STATE: dict[str, Any] = {"status": "not-requested"}


@asynccontextmanager
async def lifespan(_: FastAPI):
    if os.getenv("FUNCTION_DEPLOY_ON_STARTUP", "false").lower() == "true":
        from web.function_deploy import deploy_function

        FUNCTION_STATE.update({"status": "deploying"})
        result = await asyncio.to_thread(deploy_function)
        FUNCTION_STATE.update(result)
    if os.getenv("AGENT_DEPLOY_ON_STARTUP", "false").lower() == "true":
        from agent.deploy import deploy_agent

        AGENT_STATE.update({"status": "deploying"})
        result = await asyncio.to_thread(deploy_agent)
        AGENT_STATE.update(result)
    if os.getenv("AGENT_SMOKE_ON_STARTUP", "false").lower() == "true":
        from agent.invoke import invoke_agent

        AGENT_STATE.update({"smoke_status": "running"})
        smoke = await asyncio.to_thread(
            invoke_agent,
            "Which storage accounts grew abnormally last week? Use the private deterministic tool.",
        )
        AGENT_STATE.update({"smoke_status": "passed", "smoke": smoke})
    async with MCP_SERVER.session_manager.run():
        scheduler_task = asyncio.create_task(_discovery_scheduler())
        try:
            yield
        finally:
            scheduler_task.cancel()
            try:
                await scheduler_task
            except asyncio.CancelledError:
                pass

app = FastAPI(
    title="Storage Intelligence Agent",
    version="0.1.0",
    docs_url="/api/docs",
    openapi_url="/api/openapi.json",
    lifespan=lifespan,
)
app.mount("/static", StaticFiles(directory=STATIC), name="static")


class QueryRequest(BaseModel):
    question: str = Field(min_length=3, max_length=500)
    filters: dict[str, str] = Field(default_factory=dict)


class SavedQuestionRequest(BaseModel):
    question: str = Field(min_length=3, max_length=500)


class AddStorageAccountRequest(BaseModel):
    name: str = Field(min_length=3, max_length=24, pattern=r"^[a-z0-9]+$")
    tenant_id: str = Field(
        min_length=36,
        max_length=36,
        pattern=r"^[0-9a-fA-F]{8}(?:-[0-9a-fA-F]{4}){3}-[0-9a-fA-F]{12}$",
    )
    management_group: str = Field(min_length=1, max_length=100)
    subscription: str = Field(min_length=1, max_length=100)
    environment: str = Field(min_length=2, max_length=20)
    subsidiary: str = Field(
        min_length=1,
        max_length=100,
        validation_alias=AliasChoices("subsidiary", "business_unit"),
    )
    region: str = Field(min_length=1, max_length=50)
    tier: str = Field(min_length=1, max_length=20)


class CatalogValueRequest(BaseModel):
    value: str = Field(min_length=1, max_length=100)


class DiscoveryScheduleRequest(BaseModel):
    cron: str = Field(min_length=9, max_length=100)


class SavingsSimulationRequest(BaseModel):
    adoption_pct: int = Field(ge=1, le=100)
    filters: dict[str, str] = Field(default_factory=dict)


class ProjectOwnerNotificationRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")

    account_ids: list[Annotated[str, Field(min_length=1, max_length=512)]] = Field(
        min_length=1,
        max_length=100,
    )

    @field_validator("account_ids")
    @classmethod
    def require_unique_account_ids(cls, value: list[str]) -> list[str]:
        if len(value) != len(set(value)):
            raise ValueError("account_ids must be unique")
        return value


REQUIRED_IMPORT_COLUMNS = {
    "name",
    "tenant_id",
    "management_group",
    "subscription",
    "environment",
    "subsidiary",
    "region",
    "tier",
}
IMPORT_COLUMN_ALIASES = {
    "account": "name",
    "account_name": "name",
    "storage_account": "name",
    "storage_account_name": "name",
    "tenant": "tenant_id",
    "tenantid": "tenant_id",
    "managementgroup": "management_group",
    "management_group_id": "management_group",
    "subscription_name": "subscription",
    "deployment_environment": "environment",
    "stage": "environment",
    "env": "environment",
    "businessunit": "subsidiary",
    "business_unit": "subsidiary",
    "business_unit_name": "subsidiary",
    "subsidiary_name": "subsidiary",
    "access_tier": "tier",
    "tiers": "tier",
    "azure_region": "region",
}
MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 10_000


def _catalog_choices() -> dict[str, set[str]]:
    return {
        "tenant_id": set(TENANT_CATALOG),
        "management_group": set(MANAGEMENT_GROUP_CATALOG),
        "subscription": set(SUBSCRIPTION_CATALOG),
        "environment": set(ENVIRONMENT_CATALOG),
        "subsidiary": set(SUBSIDIARY_CATALOG),
        "region": set(AZURE_REGION_LABELS),
        "tier": {row["tier"] for row in ACCOUNTS},
    }


def _canonicalize_account(payload: AddStorageAccountRequest) -> AddStorageAccountRequest:
    updates: dict[str, str] = {"name": payload.name.lower()}
    for field, choices in _catalog_choices().items():
        value = getattr(payload, field)
        canonical = next((choice for choice in choices if choice.casefold() == value.casefold()), None)
        if canonical is None:
            raise ValueError(f"Unsupported {field}: {value}")
        updates[field] = canonical
    canonical_payload = payload.model_copy(update=updates)
    known_group = MANAGEMENT_GROUP_BY_ID.get(canonical_payload.management_group)
    if known_group and known_group["tenant_id"] != canonical_payload.tenant_id:
        raise ValueError(
            f"Management group {canonical_payload.management_group} does not belong to tenant "
            f"{canonical_payload.tenant_id}"
        )
    known_subscription = SUBSCRIPTION_BY_NAME.get(canonical_payload.subscription)
    if known_subscription and (
        known_subscription["tenant_id"] != canonical_payload.tenant_id
        or known_subscription["management_group"] != canonical_payload.management_group
        or known_subscription["subsidiary"] != canonical_payload.subsidiary
        or known_subscription["environment"] != canonical_payload.environment
    ):
        raise ValueError(
            f"Subscription {canonical_payload.subscription} does not match the selected "
            "tenant, management group, and subsidiary"
        )
    return canonical_payload


def _build_account(payload: AddStorageAccountRequest, source: str, data_as_of: str) -> dict[str, Any]:
    subscription_id = str(uuid.uuid5(uuid.NAMESPACE_URL, f"{payload.tenant_id}:{payload.subscription}"))
    return {
        "account_id": (
            f"/subscriptions/{subscription_id}/resourceGroups/"
            f"rg-{re.sub(r'[^a-z0-9]', '-', payload.subsidiary.lower()).strip('-')}/providers/Microsoft.Storage/"
            f"storageAccounts/{payload.name}"
        ),
        "name": payload.name,
        "tenant_id": payload.tenant_id,
        "management_group": payload.management_group,
        "subscription_id": subscription_id,
        "subscription": payload.subscription,
        "environment": payload.environment,
        "subsidiary": payload.subsidiary,
        "business_unit": payload.subsidiary,
        "region": payload.region,
        "tier": payload.tier,
        "capacity_tb": 0.0,
        "growth_30d_pct": 0.0,
        "monthly_cost_usd": 0.0,
        "cost_change_pct": 0.0,
        "transactions_m": 0.0,
        "cold_fraction": 0.0,
        "last_access_p90_days": 0,
        "replication": "ZRS",
        "uses_sas_keys": None,
        "shared_key_access_enabled": None,
        "public_network_access": None,
        "blob_public_access_enabled": None,
        "private_endpoint_enabled": None,
        "service_principal_access_enabled": None,
        "managed_identity_enabled": None,
        "network_security_group": None,
        "application_security_group": None,
        "project_name": None,
        "tag_business_unit": payload.subsidiary,
        "last_accessed_date": None,
        "project_defunct": None,
        "business_criticality": "medium",
        "lifecycle_policy": False,
        "inventory_age_hours": 0,
        "metrics_age_hours": 0,
        "throttling_pct": 0.0,
        "latency_ms": 0.0,
        "databricks_workspace": None,
        "fabric_lakehouse": None,
        "sap_system": None,
        "azure_data_factory": None,
        "hns_enabled": None,
        "sftp_enabled": None,
        "application_insights_resource": None,
        "azure_function_app": None,
        "log_analytics_workspace": None,
        "databricks_io_tb": 0.0,
        "small_file_ratio": 0.0,
        "version_overhead_pct": 0.0,
        "soft_delete_overhead_pct": 0.0,
        "snapshot_overhead_pct": 0.0,
        "data_as_of": data_as_of,
        "source": source,
    }


def _normalize_import_column(value: Any) -> str:
    normalized = "_".join(
        part for part in "".join(character.lower() if character.isalnum() else "_" for character in str(value)).split("_") if part
    )
    return IMPORT_COLUMN_ALIASES.get(normalized, normalized)


def _matrix_to_rows(matrix: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    if not matrix:
        raise HTTPException(status_code=422, detail="The spreadsheet is empty")
    headers = [_normalize_import_column(value) for value in matrix[0]]
    missing = sorted(REQUIRED_IMPORT_COLUMNS - set(headers))
    if missing:
        raise HTTPException(status_code=422, detail=f"Missing required columns: {', '.join(missing)}")
    rows = []
    for values in matrix[1:]:
        if not any(value is not None and str(value).strip() for value in values):
            continue
        rows.append({header: values[index] if index < len(values) else None for index, header in enumerate(headers)})
    if not rows:
        raise HTTPException(status_code=422, detail="The spreadsheet has no storage account rows")
    if len(rows) > MAX_IMPORT_ROWS:
        raise HTTPException(status_code=413, detail=f"Spreadsheet exceeds the {MAX_IMPORT_ROWS:,}-row limit")
    return rows


def _parse_import(filename: str, content: bytes) -> list[dict[str, Any]]:
    lower_name = filename.lower()
    if lower_name.endswith(".csv"):
        try:
            matrix = list(csv.reader(StringIO(content.decode("utf-8-sig"))))
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=422, detail="CSV files must use UTF-8 encoding") from exc
        return _matrix_to_rows([tuple(row) for row in matrix])
    if lower_name.endswith(".xlsx"):
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:
            raise HTTPException(status_code=422, detail="The XLSX workbook could not be read") from exc
        try:
            return _matrix_to_rows(list(workbook.active.iter_rows(values_only=True)))
        finally:
            workbook.close()
    raise HTTPException(status_code=415, detail="Upload an .xlsx or .csv spreadsheet")


def _principal(
    request: Request,
    client_principal: Annotated[str | None, Header(alias="x-ms-client-principal")] = None,
) -> dict[str, Any]:
    if os.getenv("AUTH_DISABLED", "false").lower() == "true":
        return {"identityProvider": "local", "userId": "local-developer"}
    if not client_principal:
        raise HTTPException(status_code=401, detail="Microsoft Entra authentication is required")
    try:
        return json.loads(base64.b64decode(client_principal))
    except (ValueError, json.JSONDecodeError) as exc:
        raise HTTPException(status_code=401, detail="Invalid authenticated principal") from exc


def _has_admin_role(principal: dict[str, Any]) -> bool:
    if principal.get("identityProvider") == "local":
        return True
    required_role = os.getenv("ADMIN_ROLE", "StorageIntelligence.Admin").casefold()
    roles = {
        str(role).casefold()
        for role in (principal.get("roles") or []) + (principal.get("userRoles") or [])
    }
    for claim in principal.get("claims", []):
        claim_type = str(claim.get("typ") or claim.get("type") or "").casefold()
        if claim_type.endswith("/role") or claim_type in {"role", "roles"}:
            roles.add(str(claim.get("val") or claim.get("value") or "").casefold())
    return required_role in roles


def _admin_principal(
    principal: Annotated[dict[str, Any], Depends(_principal)],
) -> dict[str, Any]:
    if not _has_admin_role(principal):
        raise HTTPException(status_code=403, detail="Storage Intelligence administrator role is required")
    return principal


def _filters(
    tenant_id: str | None = None,
    management_group: str | None = None,
    subsidiary: str | None = None,
    subscription: str | None = None,
    business_unit: str | None = None,
    region: str | None = None,
    tier: str | None = None,
    databricks: str | None = None,
    environment: str | None = None,
) -> dict[str, str]:
    return {
        key: value
        for key, value in {
            "tenant_id": tenant_id,
            "management_group": management_group,
            "subsidiary": subsidiary or business_unit,
            "subscription": subscription,
            "region": region,
            "tier": tier,
            "databricks": databricks,
            "environment": environment,
        }.items()
        if value is not None
    }


def _scope(rows: list[dict[str, Any]], filters: dict[str, str]) -> dict[str, Any]:
    return {
        "filters": filters,
        "account_count": len(rows),
        "tenant_count": len({row["tenant_id"] for row in rows}),
        "management_group_count": len({row["management_group"] for row in rows}),
        "subsidiary_count": len({row["subsidiary"] for row in rows}),
        "subscription_count": len({row["subscription"] for row in rows}),
        "environment_count": len({row["environment"] for row in rows}),
    }


def _hierarchy_projection(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "tenant_id": item.get("tenant_id"),
        "management_group": item.get("management_group"),
        "subsidiary": item.get("subsidiary") or item.get("business_unit"),
        "subscription": item.get("subscription"),
        "environment": item.get("environment"),
    }


POSTURE_FACTOR_LABELS = {
    "sas-key": "SAS Key",
    "public-access": "Public Access",
    "no-private-endpoint": "No Private Endpoint",
    "no-service-principal": "No Service Principal",
    "managed-identity": "Managed Identity",
    "no-grs-gzrs": "No GRS/GZRS",
    "nsg-asg-linked": "NSG/ASG linked",
    "defunct-projects": "Defunct Projects",
    "sftp-enabled": "SFTP Enabled",
    "app-insights-data": "AppInsights Data",
    "stale-accounts": "Stale accounts",
    "missing-lifecycle": "Missing lifecycle",
}


def _matches_posture_factor(row: dict[str, Any], factor: str) -> bool:
    if factor == "sas-key":
        return row.get("uses_sas_keys") is True
    if factor == "public-access":
        return bool(row.get("public_network_access") or row.get("blob_public_access_enabled"))
    if factor == "no-private-endpoint":
        return row.get("private_endpoint_enabled") is False
    if factor == "no-service-principal":
        return row.get("service_principal_access_enabled") is False
    if factor == "managed-identity":
        return row.get("managed_identity_enabled") is True
    if factor == "no-grs-gzrs":
        return row.get("replication") not in {"GRS", "GZRS"}
    if factor == "nsg-asg-linked":
        return bool(row.get("network_security_group") or row.get("application_security_group"))
    if factor == "defunct-projects":
        return row.get("project_defunct") is True
    if factor == "sftp-enabled":
        return row.get("sftp_enabled") is True
    if factor == "app-insights-data":
        return bool(row.get("application_insights_resource"))
    if factor == "stale-accounts":
        return row["inventory_age_hours"] > 48 or row["metrics_age_hours"] > 24
    if factor == "missing-lifecycle":
        return not row["lifecycle_policy"]
    return False


def _posture_detail(row: dict[str, Any], factor: str) -> str:
    if factor == "sas-key":
        return (
            "SAS usage observed"
            + ("; shared-key access enabled" if row.get("shared_key_access_enabled") else "")
        )
    if factor == "public-access":
        values = []
        if row.get("public_network_access"):
            values.append("public network")
        if row.get("blob_public_access_enabled"):
            values.append("blob public access")
        return " and ".join(values) + " enabled"
    if factor == "no-private-endpoint":
        return "No approved private endpoint is enabled"
    if factor == "no-service-principal":
        return "No service-principal-based access marker is enabled"
    if factor == "managed-identity":
        return "System-assigned or user-assigned managed identity is enabled"
    if factor == "no-grs-gzrs":
        return f"Replication is {row.get('replication')}; GRS/GZRS is not enabled"
    if factor == "nsg-asg-linked":
        groups = [
            value
            for value in (
                row.get("network_security_group"),
                row.get("application_security_group"),
            )
            if value
        ]
        return f"Associated with {', '.join(groups)}"
    if factor == "sftp-enabled":
        return (
            "SFTP endpoint is enabled"
            + (" with hierarchical namespace" if row.get("hns_enabled") else "")
        )
    if factor == "app-insights-data":
        return f"Stores Application Insights data for {row.get('application_insights_resource')}"
    if factor == "stale-accounts":
        return (
            f"Inventory age {row['inventory_age_hours']}h; metrics age "
            f"{row['metrics_age_hours']}h"
        )
    if factor == "missing-lifecycle":
        return "No lifecycle management policy is configured"
    return (
        f"Project {row.get('project_name') or 'Unassigned'} is defunct; "
        f"last accessed {row.get('last_accessed_date') or 'unknown'}"
    )


@app.get("/")
def index() -> FileResponse:
    return FileResponse(STATIC / "index.html")


@app.get("/healthz")
def health() -> dict[str, Any]:
    return {"status": "healthy", "service": "storage-intelligence-web"}


@app.get("/readyz")
def readiness() -> dict[str, Any]:
    return {
        "status": "ready",
        "dataset_accounts": len(ACCOUNTS),
        "dataset_fingerprint": dataset_fingerprint(ACCOUNTS),
        "connectors": "synthetic",
        "function_deployment": FUNCTION_STATE,
        "agent": AGENT_STATE,
    }


@app.get("/api/portfolio")
def get_portfolio(
    principal: Annotated[dict[str, Any], Depends(_principal)],
    tenant_id: str | None = None,
    management_group: str | None = None,
    subsidiary: str | None = None,
    subscription: str | None = None,
    business_unit: str | None = None,
    region: str | None = None,
    tier: str | None = None,
    databricks: str | None = None,
    environment: str | None = None,
) -> dict[str, Any]:
    filters = _filters(
        tenant_id,
        management_group,
        subsidiary,
        subscription,
        business_unit,
        region,
        tier,
        databricks,
        environment,
    )
    rows = ENGINE.filter(filters)
    platform_accounts = sorted(
        [
            {
                "account_id": row["account_id"],
                "name": row["name"],
                "tenant_id": row["tenant_id"],
                "management_group": row["management_group"],
                "subscription": row["subscription"],
                "environment": row["environment"],
                "subsidiary": row["subsidiary"],
                "region": row["region"],
                "tier": row["tier"],
                "databricks_workspace": row.get("databricks_workspace"),
                "fabric_lakehouse": row.get("fabric_lakehouse"),
                "sap_system": row.get("sap_system"),
                "azure_data_factory": row.get("azure_data_factory"),
                "sftp_enabled": row.get("sftp_enabled"),
                "application_insights_resource": row.get("application_insights_resource"),
                "azure_function_app": row.get("azure_function_app"),
                "log_analytics_workspace": row.get("log_analytics_workspace"),
                "managed_identity_enabled": row.get("managed_identity_enabled"),
            }
            for row in rows
        ],
        key=lambda account: account["name"],
    )
    return {
        "summary": portfolio_summary(rows),
        "risks": risks(rows, limit=None, minimum_score=AT_RISK_THRESHOLD),
        "risk_threshold": AT_RISK_THRESHOLD,
        "filters": {
            "tenant_ids": sorted(TENANT_CATALOG),
            "management_groups": sorted(MANAGEMENT_GROUP_CATALOG),
            "subscriptions": sorted(SUBSCRIPTION_CATALOG),
            "environments": sorted(ENVIRONMENT_CATALOG),
            "subsidiaries": sorted(SUBSIDIARY_CATALOG),
            "business_units": sorted(SUBSIDIARY_CATALOG),
            "regions": sorted(AZURE_REGION_LABELS, key=lambda region: AZURE_REGION_LABELS[region]),
            "tiers": sorted({row["tier"] for row in ACCOUNTS}),
        },
        "catalog": {
            "tenant_labels": {
                tenant_id: TENANT_LABELS.get(tenant_id, tenant_id)
                for tenant_id in TENANT_CATALOG
            },
            "management_group_labels": {
                group_id: MANAGEMENT_GROUP_LABELS.get(group_id, group_id)
                for group_id in MANAGEMENT_GROUP_CATALOG
            },
            "region_labels": AZURE_REGION_LABELS,
            "tracked_regions": sorted(REGION_CATALOG),
        },
        "platform_accounts": platform_accounts,
        "hierarchy": {
            "tenant_count": len({row["tenant_id"] for row in rows}),
            "management_group_count": len({row["management_group"] for row in rows}),
            "subsidiary_count": len({row["subsidiary"] for row in rows}),
            "subscription_count": len({row["subscription"] for row in rows}),
            "environment_count": len({row["environment"] for row in rows}),
        },
        "permissions": {"admin": _has_admin_role(principal)},
        "scope": _scope(rows, filters),
        "data_as_of": max(row["data_as_of"] for row in rows) if rows else None,
    }


def _run_tenant_discovery() -> None:
    try:
        result = discover_storage_accounts()
        with DISCOVERY_LOCK:
            trigger = str(DISCOVERY_STATE.get("trigger", "unknown"))
        persistence = persist_discovered_accounts(
            result["accounts"],
            pulled_at=result["pulled_at"],
            trigger=trigger,
        )
        added = 0
        updated = 0
        skipped = 0
        data_as_of = result["pulled_at"]
        with ACCOUNT_LOCK:
            by_id = {row["account_id"]: row for row in ACCOUNTS}
            by_name = {row["name"]: row for row in ACCOUNTS}
            known_tiers = {row["tier"] for row in ACCOUNTS}
            for discovered in result["accounts"]:
                region = discovered["region"]
                if region not in AZURE_REGION_LABELS:
                    skipped += 1
                    continue
                tenant_id = discovered["tenant_id"]
                management_group = discovered.get("management_group") or "Unassigned"
                subscription = discovered["subscription_name"]
                environment = discovered.get("environment") or "Unassigned"
                subsidiary = (
                    discovered.get("subsidiary")
                    or discovered.get("business_unit")
                    or "Unassigned"
                )
                tier = discovered["tier"] if discovered["tier"] in known_tiers else "Hot"
                TENANT_CATALOG.add(tenant_id)
                MANAGEMENT_GROUP_CATALOG.add(management_group)
                SUBSCRIPTION_CATALOG.add(subscription)
                ENVIRONMENT_CATALOG.add(environment)
                SUBSIDIARY_CATALOG.add(subsidiary)
                REGION_CATALOG.add(region)
                payload = AddStorageAccountRequest(
                    name=discovered["name"],
                    tenant_id=tenant_id,
                    management_group=management_group,
                    subscription=subscription,
                    environment=environment,
                    subsidiary=subsidiary,
                    region=region,
                    tier=tier,
                )
                existing = by_id.get(discovered["account_id"]) or by_name.get(discovered["name"])
                if existing is None:
                    account = _build_account(payload, "azure-cli-discovery-v1", data_as_of)
                    ACCOUNTS.append(account)
                    by_id[account["account_id"]] = account
                    by_name[account["name"]] = account
                    existing = account
                    added += 1
                else:
                    existing.update(
                        {
                            "tenant_id": tenant_id,
                            "management_group": management_group,
                            "subscription": subscription,
                            "environment": environment,
                            "subsidiary": subsidiary,
                            "business_unit": subsidiary,
                            "region": region,
                            "tier": tier,
                            "data_as_of": data_as_of,
                            "source": "azure-cli-discovery-v1",
                        }
                    )
                    updated += 1
                existing.update(
                    {
                        "account_id": discovered["account_id"],
                        "tenant_id": tenant_id,
                        "management_group": management_group,
                        "subscription_id": discovered["subscription_id"],
                        "kind": discovered["kind"],
                        "sku": discovered["sku"],
                        "tier_assumed": discovered["tier_assumed"],
                        "uses_sas_keys": discovered.get("uses_sas_keys"),
                        "shared_key_access_enabled": discovered.get("shared_key_access_enabled"),
                        "public_network_access": discovered.get("public_network_access"),
                        "blob_public_access_enabled": discovered.get("blob_public_access_enabled"),
                        "private_endpoint_enabled": discovered.get("private_endpoint_enabled"),
                        "service_principal_access_enabled": discovered.get(
                            "service_principal_access_enabled"
                        ),
                        "managed_identity_enabled": discovered.get(
                            "managed_identity_enabled"
                        ),
                        "network_security_group": discovered.get("network_security_group"),
                        "application_security_group": discovered.get("application_security_group"),
                        "project_name": discovered.get("project_name"),
                        "tag_business_unit": discovered.get("tag_business_unit") or subsidiary,
                        "last_accessed_date": discovered.get("last_accessed_date"),
                        "project_defunct": discovered.get("project_defunct"),
                        "databricks_workspace": discovered.get("databricks_workspace"),
                        "fabric_lakehouse": discovered.get("fabric_lakehouse"),
                        "sap_system": discovered.get("sap_system"),
                        "azure_data_factory": discovered.get("azure_data_factory"),
                        "hns_enabled": discovered.get("hns_enabled"),
                        "sftp_enabled": discovered.get("sftp_enabled"),
                        "application_insights_resource": discovered.get(
                            "application_insights_resource"
                        ),
                        "azure_function_app": discovered.get("azure_function_app"),
                        "log_analytics_workspace": discovered.get(
                            "log_analytics_workspace"
                        ),
                    }
                )
        with DISCOVERY_LOCK:
            DISCOVERY_STATE.update(
                {
                    "status": "completed",
                    "completed_at": datetime.now(UTC).isoformat(),
                    "tenants": len(result["tenants"]),
                    "subscriptions": result["subscriptions"],
                    "management_groups": len(result.get("management_groups") or {
                        account["management_group"] for account in result["accounts"]
                    }),
                    "subsidiaries": len(result.get("subsidiaries") or {
                        account["subsidiary"] for account in result["accounts"]
                    }),
                    "environments": len(result.get("environments") or {
                        account["environment"] for account in result["accounts"]
                    }),
                    "warnings": result.get("warnings", []),
                    "discovered": len(result["accounts"]),
                    "added": added,
                    "updated": updated,
                    "skipped": skipped,
                    "persistence": persistence["status"],
                    "persisted": persistence["upserted"],
                    "cosmos_database": persistence["database"],
                    "cosmos_container": persistence["container"],
                }
            )
    except Exception as exc:
        with DISCOVERY_LOCK:
            DISCOVERY_STATE.update(
                {
                    "status": "failed",
                    "completed_at": datetime.now(UTC).isoformat(),
                    "error": str(exc)[:1000],
                }
            )


def _begin_discovery(trigger: str) -> bool:
    with DISCOVERY_LOCK:
        if DISCOVERY_STATE["status"] == "running":
            return False
        DISCOVERY_STATE.clear()
        DISCOVERY_STATE.update(
            {
                "status": "running",
                "trigger": trigger,
                "started_at": datetime.now(UTC).isoformat(),
                "schedule": DISCOVERY_SCHEDULE["cron"],
            }
        )
        return True


def _next_discovery_run(schedule: str) -> str:
    return croniter(schedule, datetime.now(UTC)).get_next(datetime).isoformat()


def _discovery_status() -> dict[str, Any]:
    with DISCOVERY_LOCK:
        status = dict(DISCOVERY_STATE)
        status["schedule"] = DISCOVERY_SCHEDULE["cron"]
        status["next_run"] = _next_discovery_run(DISCOVERY_SCHEDULE["cron"])
        status.setdefault(
            "persistence",
            "configured"
            if os.getenv("COSMOS_INVENTORY_ENABLED", "false").lower() == "true"
            else "disabled",
        )
        status.setdefault("cosmos_database", os.getenv("COSMOS_DATABASE", "storage-intelligence"))
        status.setdefault("cosmos_container", os.getenv("COSMOS_CONTAINER", "storage-accounts"))
        return status


def _eligible_connector_records(connector_key: str, rows: list[dict[str, Any]]) -> int:
    if connector_key == "databricks":
        return sum(bool(row.get("databricks_workspace")) for row in rows)
    return len(rows)


async def _discovery_scheduler() -> None:
    last_slot: datetime | None = None
    while True:
        now = datetime.now(UTC).replace(second=0, microsecond=0)
        with DISCOVERY_LOCK:
            schedule = DISCOVERY_SCHEDULE["cron"]
        if croniter.match(schedule, now) and last_slot != now:
            last_slot = now
            if _begin_discovery("schedule"):
                asyncio.create_task(asyncio.to_thread(_run_tenant_discovery))
        await asyncio.sleep(20)


@app.post("/api/admin/discovery/pull", status_code=202)
def pull_tenant_storage_details(
    background_tasks: BackgroundTasks,
    _: Annotated[dict[str, Any], Depends(_admin_principal)],
) -> dict[str, Any]:
    if not _begin_discovery("manual"):
        raise HTTPException(status_code=409, detail="Tenant-wide storage discovery is already running")
    background_tasks.add_task(_run_tenant_discovery)
    return _discovery_status()


@app.get("/api/admin/discovery/status")
def tenant_storage_discovery_status(
    _: Annotated[dict[str, Any], Depends(_admin_principal)],
) -> dict[str, Any]:
    return _discovery_status()


@app.put("/api/admin/discovery/schedule")
def update_tenant_storage_discovery_schedule(
    payload: DiscoveryScheduleRequest,
    _: Annotated[dict[str, Any], Depends(_admin_principal)],
) -> dict[str, Any]:
    schedule = " ".join(payload.cron.split())
    if len(schedule.split()) != 5 or not croniter.is_valid(schedule):
        raise HTTPException(status_code=422, detail="Use a valid five-field cron expression")
    temporary = DISCOVERY_SCHEDULE_PATH.with_suffix(
        f"{DISCOVERY_SCHEDULE_PATH.suffix}.tmp"
    )
    try:
        DISCOVERY_SCHEDULE_PATH.parent.mkdir(parents=True, exist_ok=True)
        temporary.write_text(
            json.dumps(
                {
                    "cron": schedule,
                    "updated_at": datetime.now(UTC).isoformat(),
                },
                indent=2,
            ),
            encoding="utf-8",
        )
        temporary.replace(DISCOVERY_SCHEDULE_PATH)
    except OSError as exc:
        raise HTTPException(status_code=500, detail="Discovery schedule could not be persisted") from exc
    with DISCOVERY_LOCK:
        DISCOVERY_SCHEDULE["cron"] = schedule
        DISCOVERY_STATE["schedule"] = schedule
    return _discovery_status()


@app.post("/api/admin/connectors/{connector_key}/enable")
def enable_connector(
    connector_key: str,
    _: Annotated[dict[str, Any], Depends(_admin_principal)],
) -> dict[str, Any]:
    if connector_key not in CONNECTOR_DEFINITIONS:
        raise HTTPException(status_code=404, detail=f"Unknown connector: {connector_key}")
    with CONNECTOR_LOCK:
        runtime = CONNECTOR_RUNTIME[connector_key]
        runtime.update({"enabled": True, "status": "enabled"})
        return {
            "key": connector_key,
            "name": CONNECTOR_DEFINITIONS[connector_key]["name"],
            **runtime,
        }


@app.post("/api/admin/connectors/{connector_key}/run")
def run_connector(
    connector_key: str,
    _: Annotated[dict[str, Any], Depends(_admin_principal)],
) -> dict[str, Any]:
    if connector_key not in CONNECTOR_DEFINITIONS:
        raise HTTPException(status_code=404, detail=f"Unknown connector: {connector_key}")
    with CONNECTOR_LOCK:
        runtime = CONNECTOR_RUNTIME[connector_key]
        if not runtime["enabled"]:
            raise HTTPException(status_code=409, detail="Enable the connector before running it")
        runtime.update(
            {
                "status": "healthy",
                "records": _eligible_connector_records(connector_key, ACCOUNTS),
                "last_run": datetime.now(UTC).isoformat(),
                "mode": "pilot-fixture",
            }
        )
        return {
            "key": connector_key,
            "name": CONNECTOR_DEFINITIONS[connector_key]["name"],
            **runtime,
        }


@app.post("/api/catalog/{dimension}", status_code=201)
def add_catalog_value(
    dimension: str,
    payload: CatalogValueRequest,
    _: Annotated[dict[str, Any], Depends(_principal)],
) -> dict[str, Any]:
    value = payload.value.strip()
    if dimension == "tenants":
        if not re.fullmatch(
            r"[0-9a-fA-F]{8}(?:-[0-9a-fA-F]{4}){3}-[0-9a-fA-F]{12}",
            value,
        ):
            raise HTTPException(status_code=422, detail="Tenant ID must be a valid UUID")
        value = value.lower()
        catalog = TENANT_CATALOG
    elif dimension == "management-groups":
        if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._-]{0,99}", value):
            raise HTTPException(status_code=422, detail="Management group IDs contain unsupported characters")
        catalog = MANAGEMENT_GROUP_CATALOG
    elif dimension == "subscriptions":
        if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._ -]{0,99}", value):
            raise HTTPException(status_code=422, detail="Subscription names contain unsupported characters")
        catalog = SUBSCRIPTION_CATALOG
    elif dimension in {"business-units", "subsidiaries"}:
        if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9 &._-]{0,99}", value):
            raise HTTPException(status_code=422, detail="Subsidiary names contain unsupported characters")
        catalog = SUBSIDIARY_CATALOG
    elif dimension == "regions":
        canonical = next(
            (
                code
                for code, label in AZURE_REGION_LABELS.items()
                if code.casefold() == value.casefold() or label.casefold() == value.casefold()
            ),
            None,
        )
        if canonical is None:
            raise HTTPException(status_code=422, detail=f"Unknown Azure region: {value}")
        value = canonical
        catalog = REGION_CATALOG
    else:
        raise HTTPException(status_code=404, detail=f"Unsupported catalog dimension: {dimension}")

    with ACCOUNT_LOCK:
        existing = next((item for item in catalog if item.casefold() == value.casefold()), None)
        if existing is not None:
            raise HTTPException(status_code=409, detail=f"Value is already available: {existing}")
        catalog.add(value)
    return {
        "dimension": dimension,
        "value": value,
        "count": len(catalog),
        "note": "Added to the pilot selection catalog only; no Azure resource was created.",
    }


@app.get("/api/accounts")
def get_accounts(
    _: Annotated[dict[str, Any], Depends(_principal)],
    limit: Annotated[int, Query(ge=1, le=200)] = 50,
    offset: Annotated[int, Query(ge=0)] = 0,
    tenant_id: str | None = None,
    management_group: str | None = None,
    subsidiary: str | None = None,
    subscription: str | None = None,
    business_unit: str | None = None,
    region: str | None = None,
    tier: str | None = None,
    databricks: str | None = None,
    environment: str | None = None,
) -> dict[str, Any]:
    rows = ENGINE.filter(
        _filters(
            tenant_id,
            management_group,
            subsidiary,
            subscription,
            business_unit,
            region,
            tier,
            databricks,
            environment,
        )
    )
    return {
        "items": rows[offset : offset + limit],
        "total": len(rows),
        "offset": offset,
        "limit": limit,
        "scope": _scope(rows, _filters(
            tenant_id,
            management_group,
            subsidiary,
            subscription,
            business_unit,
            region,
            tier,
            databricks,
            environment,
        )),
    }


@app.get("/api/posture/{factor}")
def get_posture_accounts(
    factor: str,
    _: Annotated[dict[str, Any], Depends(_principal)],
    tenant_id: str | None = None,
    management_group: str | None = None,
    subsidiary: str | None = None,
    subscription: str | None = None,
    business_unit: str | None = None,
    region: str | None = None,
    tier: str | None = None,
    databricks: str | None = None,
    environment: str | None = None,
) -> dict[str, Any]:
    if factor not in POSTURE_FACTOR_LABELS:
        raise HTTPException(status_code=404, detail=f"Unknown posture factor: {factor}")
    filters = _filters(
        tenant_id,
        management_group,
        subsidiary,
        subscription,
        business_unit,
        region,
        tier,
        databricks,
        environment,
    )
    rows = ENGINE.filter(filters)
    matches = []
    for row in rows:
        if not _matches_posture_factor(row, factor):
            continue
        score = risk_score(row)
        matches.append(
            {
                **_hierarchy_projection(row),
                "account_id": row["account_id"],
                "name": row["name"],
                "region": row["region"],
                "tier": row["tier"],
                "project_name": row.get("project_name"),
                "last_accessed_date": row.get("last_accessed_date"),
                "score": score["score"],
                "risk_factors": score["risk_factors"],
                "detail": _posture_detail(row, factor),
            }
        )
    matches.sort(key=lambda item: (-item["score"], item["name"]))
    return {
        "factor": factor,
        "label": POSTURE_FACTOR_LABELS[factor],
        "count": len(matches),
        "accounts": matches,
        "scope": _scope(rows, filters),
        "data_as_of": max((row["data_as_of"] for row in rows), default=None),
    }


@app.post("/api/accounts", status_code=201)
def add_account(
    payload: AddStorageAccountRequest,
    _: Annotated[dict[str, Any], Depends(_principal)],
) -> dict[str, Any]:
    try:
        payload = _canonicalize_account(payload)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    with ACCOUNT_LOCK:
        if any(row["name"] == payload.name for row in ACCOUNTS):
            raise HTTPException(status_code=409, detail="A storage account with this name is already tracked")
        data_as_of = datetime.now(UTC).isoformat()
        account = _build_account(payload, "manual-pilot-v1", data_as_of)
        ACCOUNTS.append(account)
    return {
        "account": account,
        "total": len(ACCOUNTS),
        "note": "Added to the pilot inventory only; no Azure resource was created.",
    }


@app.post("/api/accounts/import", status_code=201)
async def import_accounts(
    spreadsheet: Annotated[UploadFile, File(description="XLSX or CSV storage account inventory")],
    _: Annotated[dict[str, Any], Depends(_principal)],
) -> dict[str, Any]:
    filename = spreadsheet.filename or ""
    content = await spreadsheet.read(MAX_IMPORT_BYTES + 1)
    await spreadsheet.close()
    if len(content) > MAX_IMPORT_BYTES:
        raise HTTPException(status_code=413, detail="Spreadsheet exceeds the 5 MB limit")
    raw_rows = _parse_import(filename, content)

    payloads: list[AddStorageAccountRequest] = []
    errors: list[str] = []
    for index, row in enumerate(raw_rows, start=2):
        try:
            candidate = AddStorageAccountRequest(
                name=str(row.get("name") or "").strip().lower(),
                tenant_id=str(row.get("tenant_id") or "").strip(),
                management_group=str(row.get("management_group") or "").strip(),
                subscription=str(row.get("subscription") or "").strip(),
                environment=str(row.get("environment") or "").strip(),
                subsidiary=str(row.get("subsidiary") or "").strip(),
                region=str(row.get("region") or "").strip(),
                tier=str(row.get("tier") or "").strip(),
            )
            payloads.append(_canonicalize_account(candidate))
        except (ValidationError, ValueError) as exc:
            errors.append(f"Row {index}: {exc}")

    duplicate_upload_names = {
        name for name in {payload.name for payload in payloads} if sum(item.name == name for item in payloads) > 1
    }
    if duplicate_upload_names:
        errors.append(f"Duplicate names in spreadsheet: {', '.join(sorted(duplicate_upload_names))}")

    with ACCOUNT_LOCK:
        existing_names = {row["name"] for row in ACCOUNTS}
        conflicts = sorted(existing_names & {payload.name for payload in payloads})
        if conflicts:
            errors.append(f"Already tracked: {', '.join(conflicts)}")
        if errors:
            raise HTTPException(status_code=422, detail={"message": "Spreadsheet validation failed", "errors": errors})
        data_as_of = datetime.now(UTC).isoformat()
        imported = [_build_account(payload, "spreadsheet-pilot-v1", data_as_of) for payload in payloads]
        ACCOUNTS.extend(imported)

    return {
        "imported": len(imported),
        "total": len(ACCOUNTS),
        "accounts": [account["name"] for account in imported],
        "note": "Added to the pilot inventory only; no Azure resources were created.",
    }


@app.post("/api/savings/simulate")
def simulate_savings(
    payload: SavingsSimulationRequest,
    _: Annotated[dict[str, Any], Depends(_principal)],
) -> dict[str, Any]:
    try:
        rows = ENGINE.filter(payload.filters)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    trust = ENGINE.answer("portfolio overview", payload.filters)
    return {
        "simulation": tier_savings(rows, payload.adoption_pct / 100, limit=20),
        "scenarios": savings_scenarios(rows),
        "scope": trust["scope"],
        "timestamp": trust["timestamp"],
        "data_as_of": trust["data_as_of"],
        "assumptions": trust["assumptions"],
        "confidence": trust["confidence"],
    }


@app.post("/api/notifications/project-owners", status_code=202)
def notify_project_owners(
    payload: ProjectOwnerNotificationRequest,
    _: Annotated[dict[str, Any], Depends(_principal)],
) -> dict[str, Any]:
    with ACCOUNT_LOCK:
        by_id = {row["account_id"]: row for row in ACCOUNTS}
        unknown = [account_id for account_id in payload.account_ids if account_id not in by_id]
        accounts = [dict(by_id[account_id]) for account_id in payload.account_ids if account_id in by_id]
    if unknown:
        raise HTTPException(
            status_code=422,
            detail={
                "message": "Unknown storage account IDs",
                "errors": unknown,
            },
        )
    try:
        return send_project_owner_notification(accounts)
    except NotificationConfigurationError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    except NotificationDeliveryError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc


@app.get("/api/findings")
def get_findings(
    _: Annotated[dict[str, Any], Depends(_principal)],
    tenant_id: str | None = None,
    management_group: str | None = None,
    subsidiary: str | None = None,
    subscription: str | None = None,
    business_unit: str | None = None,
    region: str | None = None,
    tier: str | None = None,
    databricks: str | None = None,
    environment: str | None = None,
) -> dict[str, Any]:
    filters = _filters(
        tenant_id,
        management_group,
        subsidiary,
        subscription,
        business_unit,
        region,
        tier,
        databricks,
        environment,
    )
    rows = ENGINE.filter(filters)
    findings: list[dict[str, Any]] = []
    for item in risks(rows, limit=50, minimum_score=AT_RISK_THRESHOLD):
        findings.append(
            {
                **_hierarchy_projection(item),
                "id": f"risk:{item['account_id']}",
                "category": "Risk",
                "severity": "high" if item["score"] >= 40 else "medium",
                "account_id": item["account_id"],
                "title": item["name"],
                "summary": (
                    f"Overall risk {item['score']}/100; growth {item['components']['growth']}; "
                    f"operations {item['components']['operations']}; security "
                    f"{item['components']['security']}; governance "
                    f"{item['components']['governance']}. "
                    f"Factors: {'; '.join(item['risk_factors'][:3]) or 'none'}."
                ),
                "risk_factors": item["risk_factors"],
                "project_name": item.get("project_name"),
                "last_accessed_date": item.get("last_accessed_date"),
                "value": item["score"],
            }
        )
    for item in growth_anomalies(rows, limit=25):
        findings.append(
            {
                **_hierarchy_projection(item),
                "id": f"anomaly:{item['account_id']}",
                "category": "Growth anomaly",
                "severity": "high" if abs(item["robust_z_score"]) >= 6 else "medium",
                "account_id": item["account_id"],
                "title": item["name"],
                "summary": (
                    f"30-day growth {item['growth_30d_pct']}%; robust z-score "
                    f"{item['robust_z_score']}."
                ),
                "value": abs(item["robust_z_score"]),
            }
        )
    freshness_result = freshness(rows, limit=50)
    for item in freshness_result["accounts"]:
        age = max(item["inventory_age_hours"], item["metrics_age_hours"])
        findings.append(
            {
                **_hierarchy_projection(item),
                "id": f"freshness:{item['account_id']}",
                "category": "Data freshness",
                "severity": "high" if age > 96 else "medium",
                "account_id": item["account_id"],
                "title": item["account_id"].rsplit("/", 1)[-1],
                "summary": (
                    f"Inventory age {item['inventory_age_hours']}h; metrics age "
                    f"{item['metrics_age_hours']}h. Recommended action: {item['action']}."
                ),
                "value": age,
            }
        )
    for item in top_actions(rows, limit=10):
        findings.append(
            {
                **_hierarchy_projection(item),
                "id": f"savings:{item['account_id']}",
                "category": "Savings action",
                "severity": "low",
                "account_id": item["account_id"],
                "title": item["account_id"].rsplit("/", 1)[-1],
                "summary": (
                    f"{item['action']}; modeled monthly savings "
                    f"${item['monthly_savings_usd']:,.2f}; effort {item['effort']}."
                ),
                "value": item["monthly_savings_usd"],
            }
        )
    severity_order = {"high": 0, "medium": 1, "low": 2}
    findings.sort(key=lambda item: (severity_order[item["severity"]], -item["value"]))
    counts = {
        category: sum(item["category"] == category for item in findings)
        for category in sorted({item["category"] for item in findings})
    }
    return {
        "findings": findings,
        "counts": counts,
        "total": len(findings),
        "scope": _scope(rows, filters),
        "data_as_of": max((row["data_as_of"] for row in rows), default=None),
    }


@app.get("/api/data-health")
def get_data_health(
    principal: Annotated[dict[str, Any], Depends(_principal)],
    tenant_id: str | None = None,
    management_group: str | None = None,
    subsidiary: str | None = None,
    subscription: str | None = None,
    business_unit: str | None = None,
    region: str | None = None,
    tier: str | None = None,
    databricks: str | None = None,
    environment: str | None = None,
) -> dict[str, Any]:
    filters = _filters(
        tenant_id,
        management_group,
        subsidiary,
        subscription,
        business_unit,
        region,
        tier,
        databricks,
        environment,
    )
    rows = ENGINE.filter(filters)
    freshness_result = freshness(rows, limit=200)
    stale_count = freshness_result["stale_count"]
    source_counts: dict[str, int] = {}
    for row in rows:
        source_counts[row["source"]] = source_counts.get(row["source"], 0) + 1
    sources = [
        {
            "key": "synthetic",
            "name": "Synthetic pilot",
            "status": "healthy",
            "records": source_counts.get("synthetic-v1", 0),
            "eligible_records": len(rows),
            "detail": "Deterministic seed estate",
            "can_enable": False,
            "can_run": False,
        },
        {
            "key": "azure-cli",
            "name": "Azure CLI discovery",
            "status": "healthy" if source_counts.get("azure-cli-discovery-v1", 0) else "not run",
            "records": source_counts.get("azure-cli-discovery-v1", 0),
            "eligible_records": len(rows),
            "detail": "Admin-triggered tenant inventory",
            "can_enable": False,
            "can_run": _has_admin_role(principal),
        },
    ]
    with CONNECTOR_LOCK:
        for key, definition in CONNECTOR_DEFINITIONS.items():
            runtime = CONNECTOR_RUNTIME[key]
            eligible_records = _eligible_connector_records(key, rows)
            sources.append(
                {
                    "key": key,
                    "name": definition["name"],
                    "status": runtime["status"],
                    "records": eligible_records if runtime["status"] == "healthy" else 0,
                    "eligible_records": eligible_records,
                    "detail": definition["detail"],
                    "last_run": runtime["last_run"],
                    "mode": runtime["mode"],
                    "can_enable": _has_admin_role(principal) and not runtime["enabled"],
                    "can_run": _has_admin_role(principal) and runtime["enabled"],
                }
            )
    return {
        "summary": {
            "accounts": len(rows),
            "fresh_accounts": max(0, len(rows) - stale_count),
            "stale_accounts": stale_count,
            "freshness_pct": round((len(rows) - stale_count) / max(1, len(rows)) * 100, 1),
            "missing_lifecycle_policy": sum(not row["lifecycle_policy"] for row in rows),
            "assumed_tier": sum(bool(row.get("tier_assumed")) for row in rows),
            "sas_key_accounts": sum(bool(row.get("uses_sas_keys")) for row in rows),
            "public_access_accounts": sum(
                bool(row.get("public_network_access") or row.get("blob_public_access_enabled"))
                for row in rows
            ),
            "missing_private_endpoint_accounts": sum(
                row.get("private_endpoint_enabled") is False for row in rows
            ),
            "missing_service_principal_access_accounts": sum(
                row.get("service_principal_access_enabled") is False for row in rows
            ),
            "managed_identity_accounts": sum(
                row.get("managed_identity_enabled") is True for row in rows
            ),
            "non_geo_redundant_accounts": sum(
                row.get("replication") not in {"GRS", "GZRS"} for row in rows
            ),
            "nsg_asg_linked_accounts": sum(
                bool(row.get("network_security_group") or row.get("application_security_group"))
                for row in rows
            ),
            "defunct_project_accounts": sum(
                row.get("project_defunct") is True for row in rows
            ),
            "sftp_enabled_accounts": sum(
                row.get("sftp_enabled") is True for row in rows
            ),
            "application_insights_accounts": sum(
                bool(row.get("application_insights_resource")) for row in rows
            ),
            "missing_last_access_tag": sum(
                not row.get("last_accessed_date") for row in rows
            ),
        },
        "sources": sources,
        "stale_accounts": freshness_result["accounts"],
        "permissions": {"admin": _has_admin_role(principal)},
        "scope": _scope(rows, filters),
        "data_as_of": max((row["data_as_of"] for row in rows), default=None),
    }


@app.post("/api/query")
def query(
    payload: QueryRequest,
    _: Annotated[dict[str, Any], Depends(_principal)],
) -> dict[str, Any]:
    try:
        return ENGINE.answer(payload.question, payload.filters)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/api/questions")
def get_saved_questions(
    _: Annotated[dict[str, Any], Depends(_principal)],
) -> dict[str, Any]:
    try:
        questions = list_questions()
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    return {
        "questions": questions,
        "defaults": sum(not item["custom"] for item in questions),
        "custom": sum(bool(item["custom"]) for item in questions),
    }


@app.post("/api/questions", status_code=201)
def add_saved_question(
    payload: SavedQuestionRequest,
    principal: Annotated[dict[str, Any], Depends(_principal)],
) -> dict[str, Any]:
    created_by = str(principal.get("userId") or principal.get("userDetails") or "authenticated-user")
    try:
        saved = save_question(payload.question, created_by)
        questions = list_questions()
    except (QuestionAlreadyExistsError, QuestionLibraryFullError) as exc:
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    return {
        "saved": saved,
        "questions": questions,
        "custom": sum(bool(item["custom"]) for item in questions),
    }


# Register protocol routes after the application API. The A2A SDK includes a
# tenant-aware fallback mount, so MCP must be inserted before that fallback.
app.mount("/mcp", build_mcp_http_app(MCP_SERVER), name="mcp")
register_a2a_routes(app, AGENT_SERVICE)
